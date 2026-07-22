# -*- coding: utf-8 -*-
"""اختبار كشف التسكين: السعة، التوزيع، وقواعد العائلة والجنس والاشتراك."""
import sys, io
import os as _os
_HERE = _os.path.dirname(_os.path.abspath(__file__))
_ROOT = _os.path.dirname(_HERE)
_OUTDIR = _os.path.join(_HERE, "_out")
_os.makedirs(_OUTDIR, exist_ok=True)
sys.path.insert(0, _ROOT)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from hajj_app.rooming import (
    build_rooming_plan, group_records_by_room, is_shared, room_capacity,
    room_number_in_type,
)
from hajj_app.excel_io import export_grouped_excel
from hajj_app.pdf_io import export_pdf
from hajj_app.mrz import PassportData


def rec(fam, name, sex, rtype, hotel="كونراد", num="", phone=""):
    r = PassportData()
    r.family_number = fam; r.full_name_ar = name; r.sex = sex
    r.room_type = rtype; r.hotel = hotel; r.room_number = num; r.phone = phone
    return r


print("=== السعة من قيم حقيقية غير موحّدة ===")
caps = {
    "ثلاثي": 3, "ثلاثية": 3, "ثنائي": 2, "ثنائية": 2, "ثنائى": 2,
    "رباعي": 4, "رباعية": 4, "رباعي - مشترك": 4, "رباعية مشتركة": 4,
    "مفرد": 1, "Double": 2, "Quad": 4, "Triple": 3, "خماسية": 5,
    "بدون": 4, "": 4, "رباعية 9": 4,
}
for rt, expected in caps.items():
    got = room_capacity(rt)
    assert got == expected, f"{rt!r}: توقعنا {expected}، جاء {got}"
    print(f"  OK  {rt:<16} -> {got}")

assert is_shared("رباعي - مشترك") and is_shared("رباعية مشتركة")
assert not is_shared("رباعي") and not is_shared("ثنائية")
print("  OK: كشف الاشتراك")

print("\n=== استخلاص رقم الغرفة من نوعها ===")
for rt, expected in [("رباعي 2", "2"), ("ثلاثي 3", "3"), ("رباعية 9", "9"),
                     ("رباعي", ""), ("رباعي - مشترك", ""), ("Quad", "")]:
    got = room_number_in_type(rt)
    assert got == expected, f"{rt!r}: توقعنا {expected!r}، جاء {got!r}"
    print(f"  OK  {rt:<16} -> رقم {got!r}")

print("\n=== الاعتماد على نوع الغرفة: (رباعي 2) كلهم في غرفة واحدة ===")
records = [
    rec("11", "أ", "ذكر", "رباعي 2"),
    rec("12", "ب", "ذكر", "رباعي 2"),   # عائلة مختلفة، نفس الغرفة 2
    rec("13", "ج", "ذكر", "رباعي 2"),
    rec("14", "د", "ذكر", "رباعي 2"),
    rec("21", "هـ", "أنثى", "ثلاثي 3"),
    rec("22", "و", "أنثى", "ثلاثي 3"),
    rec("23", "ز", "أنثى", "ثلاثي 3"),
]
plan = build_rooming_plan(records)
room2 = [r for r in plan.rooms if r.number == "2"]
room3 = [r for r in plan.rooms if r.number == "3"]
assert len(room2) == 1 and room2[0].count == 4, "لم يُجمع (رباعي 2) في غرفة واحدة"
assert room2[0].capacity == 4 and not room2[0].auto_numbered
assert len(room3) == 1 and room3[0].count == 3, "لم يُجمع (ثلاثي 3) في غرفة واحدة"
assert room3[0].capacity == 3
print("  OK: (رباعي 2) -> غرفة 2 فيها 4، (ثلاثي 3) -> غرفة 3 فيها 3")

print("\n=== رباعي 2 و رباعي 3 غرفتان مختلفتان ===")
records = [
    rec("1", "أ", "ذكر", "رباعي 2"),
    rec("1", "ب", "ذكر", "رباعي 2"),
    rec("2", "ج", "ذكر", "رباعي 3"),
]
plan = build_rooming_plan(records)
assert {r.number for r in plan.rooms} == {"2", "3"}, [r.number for r in plan.rooms]
print("  OK: نفس النوع برقمين مختلفين -> غرفتان")

print("\n=== نفس الرقم بأنواع مختلفة = غرف مختلفة ===")
# مفرد 1 و ثنائية 1 و رباعية 1: رقمها واحد لكنها ثلاث غرف مختلفة
records = [
    rec("1", "أ", "ذكر", "مفرد 1"),
    rec("2", "ب", "ذكر", "ثنائية 1"),
    rec("2", "ج", "ذكر", "ثنائية 1"),
    rec("3", "د", "ذكر", "رباعية 1"),
]
plan = build_rooming_plan(records)
assert len(plan.rooms) == 3, f"دُمجت غرف مختلفة الأنواع: {len(plan.rooms)}"
caps = sorted(r.capacity for r in plan.rooms)
assert caps == [1, 2, 4], caps
print("  OK: مفرد 1، ثنائية 1، رباعية 1 -> ثلاث غرف منفصلة")

print("\n=== الترتيب: مفرد ثم ثنائي ثم ثلاثي ثم رباعي، كلٌّ بأرقامه ===")
records = [
    rec("1", "ح", "ذكر", "رباعية 2"), rec("1", "ط", "ذكر", "رباعية 2"),
    rec("2", "أ", "ذكر", "مفرد 2"),
    rec("3", "ي", "ذكر", "رباعية 1"), rec("3", "ك", "ذكر", "رباعية 1"),
    rec("4", "ب", "ذكر", "مفرد 1"),
    rec("5", "ج", "ذكر", "ثلاثية 1"), rec("5", "د", "ذكر", "ثلاثية 1"),
    rec("6", "و", "ذكر", "ثنائية 2"), rec("6", "ز", "ذكر", "ثنائية 2"),
    rec("7", "هـ", "ذكر", "ثنائية 1"), rec("7", "ل", "ذكر", "ثنائية 1"),
]
plan = build_rooming_plan(records)
order = [(r.capacity, r.number) for r in plan.rooms]
expected = [(1, "1"), (1, "2"), (2, "1"), (2, "2"), (3, "1"), (4, "1"), (4, "2")]
assert order == expected, f"\nترتيب خاطئ: {order}\nالمتوقع:   {expected}"
print("  OK: " + " ← ".join(f"{'مفرد ثنائي ثلاثي رباعي'.split()[c-1]} {n}"
                            for c, n in order))

print("\n=== خانة رقم الغرفة تتقدّم على الرقم المدمج ===")
records = [rec("1", "أ", "ذكر", "رباعي 2", num="500")]
plan = build_rooming_plan(records)
assert plan.rooms[0].number == "500", plan.rooms[0].number
print("  OK: خانة 'رقم الغرفة'=500 تتقدّم على '2' المدمجة")

print("\n=== العائلة تبقى معاً ===")
records = [rec("1", f"فرد {i}", "ذكر", "رباعي") for i in range(1, 5)]
plan = build_rooming_plan(records)
assert len(plan.rooms) == 1 and plan.rooms[0].count == 4
assert plan.rooms[0].families == ["1"]
print("  OK: عائلة من 4 في رباعية واحدة")

print("\n=== عائلة أكبر من السعة تمتد لغرفتين ===")
records = [rec("1", f"فرد {i}", "ذكر", "ثنائي") for i in range(1, 6)]  # 5 في ثنائي
plan = build_rooming_plan(records)
assert len(plan.rooms) == 3, [r.count for r in plan.rooms]
assert sorted(r.count for r in plan.rooms) == [1, 2, 2]
assert all(r.families == ["1"] for r in plan.rooms)
print(f"  OK: 5 أفراد ثنائي -> {sorted(r.count for r in plan.rooms)} غرف، كلها نفس العائلة")

print("\n=== الاشتراك يُعبَّر عنه بنفس رقم الغرفة ===")
# عائلتان مختلفتان بنفس رقم الغرفة -> غرفة واحدة (هذا هو الاشتراك الآن)
records = [
    rec("1", "ذكر أ", "ذكر", "رباعي 7"),
    rec("1", "ذكر ب", "ذكر", "رباعي 7"),
    rec("2", "ذكر ج", "ذكر", "رباعي 7"),
    rec("2", "ذكر د", "ذكر", "رباعي 7"),
]
plan = build_rooming_plan(records)
assert len(plan.rooms) == 1 and plan.rooms[0].count == 4
assert set(plan.rooms[0].families) == {"1", "2"}
print("  OK: عائلتان بنفس الرقم (رباعي 7) -> غرفة واحدة")

print("\n=== أرقام مختلفة لا تُدمج ولو نفس النوع ===")
records = [
    rec("1", "أ", "ذكر", "رباعي 1"),
    rec("2", "ب", "ذكر", "رباعي 2"),
]
plan = build_rooming_plan(records)
assert len(plan.rooms) == 2, "دُمجت غرفتان مختلفتا الرقم"
print("  OK: رباعي 1 و رباعي 2 -> غرفتان")

print("\n=== نوع بلا رقم: كل عائلة في غرفتها، بأرقام تلقائية ===")
records = [
    rec("1", "أ", "ذكر", "رباعي"),
    rec("1", "ب", "ذكر", "رباعي"),
    rec("2", "ج", "ذكر", "رباعي"),        # عائلة أخرى -> غرفة أخرى
]
plan = build_rooming_plan(records)
assert len(plan.rooms) == 2, "خُلطت عائلتان في نوع بلا رقم"
assert all(r.auto_numbered for r in plan.rooms)
print("  OK: عائلتان بنوع بلا رقم -> غرفتان تلقائيتان منفصلتان")

print("\n=== الفنادق منفصلة ===")
records = [
    rec("1", "أ", "ذكر", "رباعي 2", hotel="كونراد"),
    rec("2", "ب", "ذكر", "رباعي 2", hotel="هيلتون"),
]
plan = build_rooming_plan(records)
assert len(plan.rooms) == 2, "دُمج ساكنان من فندقين مختلفين"
hotels = {r.hotel for r in plan.rooms}
assert hotels == {"كونراد", "هيلتون"}
print("  OK: لا يُدمج سكان فندقين مختلفين")

print("\n=== من بلا نوع غرفة لا يُسكّن، ويُبلَّغ عنه ===")
records = [rec("1", "مسكّن", "ذكر", "ثنائي"), rec("2", "بلا سكن", "ذكر", "")]
plan = build_rooming_plan(records)
assert len(plan.unplaced) == 1 and plan.unplaced[0].name == "بلا سكن"
assert any("بلا نوع غرفة" in n for n in plan.notes)
print("  OK: الحاج بلا نوع غرفة مستبعَد ومُبلَّغ عنه")

print("\n=== تجاوز السعة يُرصد كتنبيه ===")
records = [rec("1", f"فرد {i}", "ذكر", "ثنائي", num="10") for i in range(1, 5)]  # 4 في ثنائي مرقّم
plan = build_rooming_plan(records)
room = plan.rooms[0]
assert room.count == 4 and room.capacity == 2
assert any("يتجاوز السعة" in w for w in room.warnings())
print(f"  OK: غرفة 10 بها {room.count} في سعة {room.capacity} -> تنبيه")

print("\n=== كشف التسكين العام: نفس أعمدة الطباعة، مجموعاً بالغرف ===")
records = [
    rec("1", "عبدالله الشامسي", "ذكر", "رباعي 3", phone="0501234567"),
    rec("1", "محمد الشامسي", "ذكر", "رباعي 3"),
    rec("2", "مريم النيادي", "أنثى", "ثلاثي 1", hotel="هيلتون"),
]
pdf = _os.path.join(_OUTDIR, "rooming_test.pdf")
xlsx = _os.path.join(_OUTDIR, "rooming_test.xlsx")
export_pdf(records, pdf, title="كشف التسكين", group_by_room=True)
export_grouped_excel(records, xlsx, title="كشف التسكين")

assert _os.path.getsize(pdf) > 3000
with open(pdf, "rb") as fh:
    assert fh.read(5) == b"%PDF-"
print(f"  OK: PDF ({_os.path.getsize(pdf)} بايت)")

wb = load_workbook(xlsx)
ws = wb.active
# نفس أعمدة الطباعة الكاملة (PDF_FIELDS)، لا الأعمدة القليلة القديمة
from hajj_app.fields import PDF_FIELDS
header = [c.value for c in ws[1]]
assert header == [f.label for f in PDF_FIELDS], header
assert "رقم الغرفة" in header and "المواصلات" in header
assert "الهاتف المتحرك" not in header
# صفوف عناوين الغرف + صفوف السكان: 3 حجاج + سطرا عنوان غرفتين = 5 صفوف بيانات
non_empty = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]
assert len(non_empty) == 5, len(non_empty)
print(f"  OK: إكسل بأعمدة الطباعة الكاملة، مجموعاً بالغرف ({len(non_empty)} سطر)")

print("\n=== الطباعة المفصولة بالغرف (group_by_room) ===")
from hajj_app.pdf_io import _grouped_rooms, export_pdf

records = [
    rec("11", "سالم الظاهري", "ذكر", "ثنائية 2"),
    rec("11", "راشد الظاهري", "ذكر", "ثنائية 2"),
    rec("22", "عمر البدو", "ذكر", "ثنائية 1"),
    rec("22", "بدر البدو", "ذكر", "ثنائية 1"),
    rec("33", "أحمد الزامل", "ذكر", "مفرد 1"),
    rec("44", "بلا غرفة", "ذكر", ""),
]
groups = _grouped_rooms(records)   # (عنوان، سعة، سكان)
# مرتّبة: مفرد 1، ثنائية 1، ثنائية 2، ثم بلا غرفة أخيراً
labels = [lbl for lbl, _cap, _occ in groups]
assert "مفرد" in labels[0], labels
assert "بدون غرفة" in labels[-1], labels
# السعة تُعاد لتلوين النوع: مفرد=1، ثنائي=2
assert groups[0][1] == 1, groups[0]
idx1 = next(i for i, l in enumerate(labels) if "ثنائية 1" in l)
idx2 = next(i for i, l in enumerate(labels) if "ثنائية 2" in l)
assert idx1 < idx2, labels
assert groups[idx1][1] == 2, "سعة الثنائية يجب أن تكون 2 للتلوين"
# كل غرفة ثنائية تضم شخصين، ولا يُفقد أحد
assert any(len(occ) == 2 for _l, _c, occ in groups)
assert sum(len(occ) for _l, _c, occ in groups) == 6, "فُقد أحد الحجاج في التجميع"
print("  OK: " + " | ".join(f"{l} ({len(o)})" for l, _c, o in groups))

pdf = _os.path.join(_OUTDIR, "grouped_test.pdf")
export_pdf(records, pdf, title="كشف — ثنائي", group_by_room=True)
assert _os.path.getsize(pdf) > 3000
with open(pdf, "rb") as fh:
    assert fh.read(5) == b"%PDF-"
print("  OK: PDF مفصول بالغرف وملوّن بالنوع صدر سليماً")

# تلوين النوع + المفتاح
from hajj_app.pdf_io import _room_legend, _room_type_color, _ROOM_TYPE_COLORS
assert _room_type_color(1) != _room_type_color(2) != _room_type_color(4)
legend = _room_legend(groups)   # فيه مفرد وثنائي على الأقل
assert legend is not None, "لم يُبنَ مفتاح الألوان"
print("  OK: لكل نوع لون مميّز، ومفتاح الألوان يُبنى")

print("\n*** ROOMING TESTS PASSED ***")

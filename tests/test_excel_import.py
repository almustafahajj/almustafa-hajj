# -*- coding: utf-8 -*-
"""حالات استيراد إكسل مأخوذة من كشوف حقيقية.

كل حالة هنا كانت خللاً فعلياً: عمود لا يُطابَق، ورقة خاطئة تُقرأ، أو صفوف
فارغة تُستورد كسجلات.
"""
import sys, io
import os as _os
_HERE = _os.path.dirname(_os.path.abspath(__file__))
_ROOT = _os.path.dirname(_HERE)
_OUTDIR = _os.path.join(_HERE, "_out")
_os.makedirs(_OUTDIR, exist_ok=True)
sys.path.insert(0, _ROOT)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import Workbook
from hajj_app.excel_io import import_excel
from hajj_app.fields import match_column

print("=== رؤوس كانت مُهمَلة في كشوف حقيقية ===")
cases = [
    ("الاسم الرباعي", "full_name_ar"),
    ("الاسم الثلاثي", "full_name_ar"),
    ("الاسم الانجليزي", "full_name_en"),
    ("الاسم الإنجليزي", "full_name_en"),
    ("FIRST NAME", "given_names_en"),
    ("LAST NAME", "surname_en"),
    ("First Name", "given_names_en"),
    ("رقم المرجعي", "reference_number"),
    ("هاتف الامارات", "phone"),
    ("PASSPORT EXPIRY (DD-MM-YYYY)", "expiry_date"),
    ("DOB (DD-MM-YYYY)", "birth_date"),
    ("الجنسية - Nationality", "nationality_ar"),
    ("رقم الرحلة", "flight_number"),
    ("درجة السفر", "travel_class"),
    ("المسؤول", "staff"),
    ("الموظف المسؤول", "staff"),
    ("الطيران", "airline"),          # يبقى منفصلاً عن رقم الرحلة
]
for header, expected in cases:
    got = match_column(header)
    assert got == expected, f"{header!r} -> {got!r}، والمتوقع {expected!r}"
    print(f"  OK  {header:<30} -> {expected}")

print("\n=== أعمدة «الوقت» المجرّدة تُحدَّد من التاريخ المجاور ===")
from openpyxl import Workbook
path = _os.path.join(_OUTDIR, "bare_times.xlsx")
wb = Workbook(); ws = wb.active
# ترتيب كملف الحملة: الوقت يسبق تاريخه
ws.append(["م", "اسم الحاج", "رقم الجواز", "الوقت", "تاريخ المغادرة",
           "الوقت", "تاريخ الوصول", "تنفيذي", "سيارة خاصة",
           "قيمة البرنامج مع الهدي"])
ws.append([1, "عمر حسن", "A111", "23:30", "2026-06-20",
           "08:15", "2026-05-18", "نعم", "باص 9", "44,000"])
wb.save(path)
recs, notes = import_excel(path)
assert len(recs) == 1, (len(recs), notes)
r = recs[0]
assert r.arrival_time == "08:15", r.arrival_time
assert r.departure_time == "23:30", r.departure_time
assert r.executive_service == "نعم", r.executive_service
assert r.transport == "باص 9", r.transport
assert r.program_value == "44,000", r.program_value
# لا شيء من هذه الأعمدة يظهر في تنبيه "لم يتم التعرّف عليها"
ignored = " ".join(notes)
for missed in ("الوقت", "تنفيذي", "سيارة خاصة", "قيمة البرنامج"):
    assert missed not in ignored, f"{missed} ما زال مُتجاهَلاً: {notes}"
print(f"  OK: الوقت-وصول={r.arrival_time}، الوقت-مغادرة={r.departure_time}، "
      f"تنفيذي/مواصلات/قيمة كلها قُرئت")

print("\n=== الاسم اللاتيني مقسوماً عمودين ===")
path = _os.path.join(_OUTDIR, "split_name.xlsx")
wb = Workbook(); ws = wb.active
ws.append(["NO", "LAST NAME", "FIRST NAME", "PASSPORT NUMBER", "DOB (DD-MM-YYYY)"])
ws.append([1, "ALSHAMSI", "ABDULLA ALI MOHAMMED", "AA0319030", "27/04/1985"])
ws.append([2, "ALBLOOSHI", "ASIA MOHAMED ROSTOM", "AA1093552", "16/01/1985"])
wb.save(path)

records, notes = import_excel(path)
assert len(records) == 2, (len(records), notes)
assert records[0].full_name_en == "ABDULLA ALI MOHAMMED ALSHAMSI", records[0].full_name_en
assert records[0].passport_number == "AA0319030"
assert records[0].birth_date == "1985-04-27", records[0].birth_date
print(f"  OK: دُمج العمودان -> {records[0].full_name_en!r}")
print(f"  OK: التاريخ ذو اللاحقة بين قوسين -> {records[0].birth_date}")

print("\n=== صفوف الترقيم الفارغة لا تصير سجلات ===")
path = _os.path.join(_OUTDIR, "empty_rows.xlsx")
wb = Workbook(); ws = wb.active
ws.append(["NO", "LAST NAME", "FIRST NAME", "PASSPORT NUMBER"])
ws.append([1, "ALSHAMSI", "ABDULLA", "AA0319030"])
for n in range(2, 30):                       # قالب فارغ: ترقيم بلا بيانات
    ws.append([n, "", "", ""])
wb.save(path)

records, notes = import_excel(path)
assert len(records) == 1, f"استُوردت صفوف فارغة: {len(records)}"
assert any("بلا اسم" in n for n in notes), notes
print(f"  OK: سجل واحد من 29 صفاً، والباقي أُبلغ عنه -> {[n for n in notes if 'بلا اسم' in n][0]}")

print("\n=== تُختار الورقة ذات البيانات لا القالب الفارغ ===")
path = _os.path.join(_OUTDIR, "many_sheets.xlsx")
wb = Workbook()
# قالب مثالي الرؤوس لكنه فارغ — كان يفوز سابقاً لكثرة أعمدته
template = wb.active; template.title = "FLIGHT"
template.append(["NO", "LAST NAME", "FIRST NAME", "PASSPORT NUMBER",
                 "DOB (DD-MM-YYYY)", "GENDER", "Nationality", "الفندق"])
for n in range(1, 60):
    template.append([n, "", "", "", "", "", "", ""])
# ورقة أقل أعمدة لكن فيها حجاج حقيقيون
real = wb.create_sheet("ROOMING LIST")
real.append(["اسم الحاج", "الفندق", "الهاتف المتحرك"])
real.append(["اليازيه حميد سالم هدفه العامرى", "كونراد", "971501111111"])
real.append(["خوله محمد حسن الظفير", "كونراد", "971502222222"])
wb.save(path)

records, notes = import_excel(path)
assert len(records) == 2, (len(records), notes)
assert records[0].full_name_ar == "اليازيه حميد سالم هدفه العامرى", records[0].full_name_ar
assert any("ROOMING LIST" in n for n in notes), notes
print(f"  OK: قُرئت الورقة الصحيحة، و{len(records)} حاجاً استُوردا")
print(f"  OK: أُبلغ المستخدم بالورقة المختارة -> {[n for n in notes if 'قُرئت' in n][0]}")

print("\n=== الاسم الكامل الموجود لا يُستبدل بالعمودين ===")
path = _os.path.join(_OUTDIR, "both_names.xlsx")
wb = Workbook(); ws = wb.active
ws.append(["الاسم الانجليزي", "FIRST NAME", "LAST NAME", "رقم الجواز"])
ws.append(["FULL NAME FROM PASSPORT", "IGNORED", "ALSO IGNORED", "AA0000001"])
wb.save(path)
records, _ = import_excel(path)
assert records[0].full_name_en == "FULL NAME FROM PASSPORT", records[0].full_name_en
print(f"  OK: {records[0].full_name_en!r} لم يُستبدل بدمج العمودين")

print("\n*** EXCEL IMPORT TESTS PASSED ***")

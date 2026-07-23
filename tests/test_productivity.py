# -*- coding: utf-8 -*-
"""اختبار الإنتاجية: محتوى QR، كشف المواصلات، وبطاقات الحجّاج."""
import sys, io
import os as _os
_HERE = _os.path.dirname(_os.path.abspath(__file__))
_ROOT = _os.path.dirname(_HERE)
_OUTDIR = _os.path.join(_HERE, "_out")
_os.makedirs(_OUTDIR, exist_ok=True)
sys.path.insert(0, _ROOT)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from hajj_app.cards import qr_payload
from hajj_app.transport import (
    TRANSPORT_COLUMNS, distinct_transports, executive_display,
    export_transport_excel, group_by_transport,
)
from hajj_app.pdf_io import export_badges_pdf, export_transport_pdf
from hajj_app.mrz import PassportData


def rec(**kw):
    return PassportData(**kw)


print("=== محتوى رمز QR للبطاقة ===")
r = rec(full_name_ar="عبدالله الشامسي", passport_number="a1234567",
        phone="0501112233", hotel="الصفوة", room_type="رباعية 2",
        family_number="101")
payload = qr_payload(r)
assert payload.startswith("الحاج: عبدالله الشامسي")
assert "الجواز: A1234567" in payload            # يُكبّر الأحرف
assert "الهاتف: 0501112233" in payload
assert "الإقامة: الصفوة - غرفة 2" in payload     # الرقم من نوع الغرفة
assert "العائلة: 101" in payload
# حاج بلا بيانات -> سطر الاسم فقط
assert qr_payload(rec()) == "الحاج: —"
print("  OK: أسطر مقروءة تُعرّف الحاج")

print("\n=== تجميع المواصلات ===")
recs = [
    rec(full_name_ar="أ", transport="باص 1", passport_number="P1", hotel="الصفوة"),
    rec(full_name_ar="ب", transport="باص 2", passport_number="P2"),
    rec(full_name_ar="ج", transport="باص 1", passport_number="P3"),
    rec(full_name_ar="د", transport="", passport_number="P4"),   # بلا مواصلات
]
assert distinct_transports(recs) == ["باص 1", "باص 2"]
groups, unassigned = group_by_transport(recs)
assert [name for name, _ in groups] == ["باص 1", "باص 2"]       # مرتّبة
assert len(groups[0][1]) == 2 and len(unassigned) == 1
print(f"  OK: باص 1 فيه 2، باص 2 فيه 1، وبلا مواصلات 1")

print("\n=== خدمة التنفيذي: القيمة كما في الكشف العام ===")
assert executive_display(rec(executive_service="جيمس")) == "جيمس"
assert executive_display(rec(executive_service="نعم")) == "نعم"       # أي قيمة تُعرض كما هي
assert executive_display(rec(executive_service=" VIP ")) == "VIP"     # تُشذّب الفراغات
assert executive_display(rec(executive_service="")) == ""
assert executive_display(rec()) == ""
print("  OK: تُعرض قيمة خدمة التنفيذي كما في الكشف العام")

print("\n=== تصدير المواصلات إكسل ===")
xlsx = _os.path.join(_OUTDIR, "transport.xlsx")
export_transport_excel(recs, xlsx)
wb = load_workbook(xlsx); ws = wb.active
assert ws.sheet_view.rightToLeft is True
assert [c.value for c in ws[2]] == list(TRANSPORT_COLUMNS)
assert "رقم الجواز" not in TRANSPORT_COLUMNS and "الغرفة" not in TRANSPORT_COLUMNS
assert list(TRANSPORT_COLUMNS) == [
    "م", "اسم الحاج", "الهاتف", "الفندق", "خدمة التنفيذي", "كرسي متحرك"]
serials = [row[0].value for row in ws.iter_rows(min_row=3) if isinstance(row[0].value, int)]
assert len(serials) == 4, serials     # كل الحجّاج مُدرجون (مع مجموعة بلا مواصلات)
print(f"  OK: إكسل RTL بأعمدة {list(TRANSPORT_COLUMNS)}، {len(serials)} حاجاً")

print("\n=== تصدير المواصلات PDF (كل باص في صفحة واحدة) ===")
pdf = _os.path.join(_OUTDIR, "transport.pdf")
export_transport_pdf(recs, pdf)
assert _os.path.getsize(pdf) > 3000
with open(pdf, "rb") as fh:
    assert fh.read(5) == b"%PDF-"
import fitz
d = fitz.open(pdf)
# مجموعتان (باص 1، باص 2) + بلا مواصلات = 3 صفحات، كل واحدة صفحة، وطولية
assert d.page_count == 3, d.page_count
page = d[0]
assert page.rect.height > page.rect.width, "الصفحة يجب أن تكون طولية"
d.close()
# باص كبير (60 راكباً بأسماء طويلة وفنادق) يبقى صفحة واحدة، كل حاج سطر واحد
big = [rec(full_name_ar=f"عبدالله محمد راشد الشامسي {i}", transport="باص 7",
           phone=f"05{i:08d}", hotel="فندق دار الصفوة المكية",
           wheelchair=("نعم" if i % 7 == 0 else ""))
       for i in range(60)]
big_pdf = _os.path.join(_OUTDIR, "transport_big.pdf")
export_transport_pdf(big, big_pdf)
d = fitz.open(big_pdf)
assert d.page_count == 1, f"باص 7 يجب أن يكون في صفحة واحدة، لا {d.page_count}"
d.close()
print(f"  OK: 3 صفحات طولية، وباص 7 (60 راكباً) في صفحة واحدة")

print("\n=== بطاقات الحجّاج (8 لكل ورقة A4 + ورقة خلفية واحدة) ===")
from hajj_app.cards import badge_name, is_woman
# اسم البطاقة: الأول والثاني والأخير فقط
assert badge_name(rec(full_name_ar="عبدالله محمد راشد الشامسي")) == "عبدالله محمد الشامسي"
assert badge_name(rec(full_name_ar="سالم أحمد")) == "سالم أحمد"
assert is_woman(rec(sex="أنثى")) and not is_woman(rec(sex="ذكر"))
import math
recs = [rec(full_name_ar=f"عبدالله محمد الشامسي {i}",
            sex=("أنثى" if i % 2 else "ذكر"), phone=f"05011122{i:02d}",
            hotel="فندق دار الصفوة") for i in range(12)]   # >10 -> صفحتا وجوه
badges = _os.path.join(_OUTDIR, "badges.pdf")
export_badges_pdf(recs, badges, company="المصطفى للحج والعمرة", session=None,
                  preacher="0555000000", admins="خالد المدير\nسعيد المشرف",
                  emergency="0509999999")
assert _os.path.getsize(badges) > 3000
with open(badges, "rb") as fh:
    assert fh.read(5) == b"%PDF-"
import fitz
d = fitz.open(badges)
# A4 عرضية: 10 بطاقات (5×2) لكل ورقة وجوه + ورقة خلفية واحدة
PER = 10
front_pages = math.ceil(len(recs) / PER)
assert d.page_count == front_pages + 1, d.page_count      # 12 -> 2 + 1 = 3
# الصفحات عرضية (842×595)
page = d[0]
assert page.rect.width > page.rect.height, "الصفحة يجب أن تكون عرضية"
assert abs(page.rect.width - 842) < 3 and abs(page.rect.height - 595) < 3, page.rect
d.close()
# قائمة فارغة لا تتعطّل
export_badges_pdf([], _os.path.join(_OUTDIR, "badges_empty.pdf"))
print(f"  OK: A4 عرضية، {front_pages} ورقة وجوه (10/ورقة) + ورقة خلفية واحدة")

print("\n=== الاستيكرات (حقائب/غرف/أظرف) ===")
from hajj_app.pdf_io import (export_stickers_pdf, STICKER_KINDS, STICKER_LABELS,
                             _sticker_items)
srecs = [rec(full_name_ar=f"حاج رقم {i}", passport_number=f"A100000{i}",
             phone=f"05011122{i:02d}", hotel="كونراد مكة",
             room_number=str(101 + i // 2), room_type="ثنائية",
             flight_number="SV553", transport="7") for i in range(6)]
# الحقائب والأظرف: استيكر لكل حاج؛ الغرف: استيكر لكل غرفة
assert len(_sticker_items(srecs, "bag", "الحملة")) == 6
assert len(_sticker_items(srecs, "envelope", "الحملة")) == 6
rooms = _sticker_items(srecs, "room", "الحملة")
assert len(rooms) == 3, len(rooms)          # 6 حجّاج في 3 غرف (2 لكل غرفة)
assert rooms[0]["big"].startswith("غرفة") and len(rooms[0]["lines"]) == 2
for k in STICKER_KINDS:
    p = _os.path.join(_OUTDIR, f"stickers_{k}.pdf")
    export_stickers_pdf(srecs, p, kind=k, company="المصطفى للحج والعمرة")
    assert _os.path.getsize(p) > 3000
    with open(p, "rb") as fh:
        assert fh.read(5) == b"%PDF-"
# نوع غير معروف يتراجع، والقائمة الفارغة لا تتعطّل
export_stickers_pdf(srecs, _os.path.join(_OUTDIR, "stickers_x.pdf"), kind="زبد")
export_stickers_pdf([], _os.path.join(_OUTDIR, "stickers_empty.pdf"), kind="room")
assert set(STICKER_LABELS) == set(STICKER_KINDS)
print("  OK: 3 أنواع استيكرات، الغرف مجمّعة، والفارغ آمن")

print("\n*** PRODUCTIVITY TESTS PASSED ***")

# -*- coding: utf-8 -*-
"""اختبار كشف الطيران بالإنجليزية وإدخالات أماديوس."""
import sys, io
import os as _os
_HERE = _os.path.dirname(_os.path.abspath(__file__))
_ROOT = _os.path.dirname(_HERE)
_OUTDIR = _os.path.join(_HERE, "_out")
_os.makedirs(_OUTDIR, exist_ok=True)
sys.path.insert(0, _ROOT)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from hajj_app.airline import (
    AIRLINE_COLUMNS, airline_rows, amadeus_entries, export_airline_excel,
    gender_code, nationality_code, split_name, travel_class_code,
)
from hajj_app.pdf_io import export_airline_pdf
from hajj_app.mrz import PassportData


def rec(**kw):
    return PassportData(**kw)


print("=== التعيينات ===")
r = rec(surname_en="ALSHAMSI", given_names_en="ABDULLA ALI MOHAMMED",
        sex="ذكر", nationality="ARE", travel_class="بزنس")
assert split_name(r) == ("ALSHAMSI", "ABDULLA ALI MOHAMMED")
assert gender_code(r) == "M"
assert nationality_code(r) == "ARE"
assert travel_class_code(r) == "J"
print("  OK: اللقب/الاسم، M، ARE، J")

# القسمة من الاسم الكامل حين يغيب MRZ
r2 = rec(full_name_en="ABDULLA ALI ALSHAMSI", sex="أنثى",
         nationality_ar="مصر", travel_class="سياحية")
assert split_name(r2) == ("ALSHAMSI", "ABDULLA ALI"), split_name(r2)
assert gender_code(r2) == "F"
assert nationality_code(r2) == "EGY", nationality_code(r2)   # من الاسم العربي
assert travel_class_code(r2) == "Y"
print("  OK: القسمة من الاسم الكامل، F، رمز من العربي، Y")

print("\n=== صف الطيران بالترتيب المطلوب ===")
r3 = rec(surname_en="ALSHAMSI", given_names_en="ABDULLA", passport_number="aa0319030",
         expiry_date="2028-07-09", birth_date="1985-04-27", sex="ذكر",
         nationality="ARE", travel_class="بزنس", family_number="101", pnr="ab12cd")
row = airline_rows([r3])[0]
assert AIRLINE_COLUMNS[0] == "#" and AIRLINE_COLUMNS[-1] == "PNR"
# #, Last, First, Passport, Expiry, DOB, Gender, Nationality, Class, Family, PNR
assert row == [1, "ALSHAMSI", "ABDULLA", "AA0319030", "09JUL28", "27APR85",
               "M", "ARE", "J", "101", "AB12CD"], row
print(f"  OK: {row}")

print("\n=== إدخالات أماديوس ===")
from hajj_app.airline import amadeus_docs, amadeus_entry, amadeus_name
assert amadeus_name(r3) == "NM1ALSHAMSI/ABDULLA MR"
assert amadeus_docs(r3) == "SR DOCS HK1-P-ARE-AA0319030-ARE-27APR85-M-09JUL28-ALSHAMSI-ABDULLA"
assert amadeus_entry(r3) == amadeus_name(r3) + "\n" + amadeus_docs(r3)
ama = amadeus_entries([r3])
print("  " + ama.replace("\n", "\n  "))
assert amadeus_name(r3) in ama and amadeus_docs(r3) in ama
print("  OK: سطر الاسم + سطر DOCS، ونسخ فردي عبر amadeus_entry")

# طفل -> MSTR/MISS
child = rec(surname_en="ALSHAMSI", given_names_en="SARA", sex="أنثى",
            birth_date="2020-01-01", nationality="ARE", passport_number="X1")
assert "NM1ALSHAMSI/SARA MISS" in amadeus_entries([child])
print("  OK: الطفلة -> MISS")

print("\n=== تصدير إكسل (LTR إنجليزي) ===")
xlsx = _os.path.join(_OUTDIR, "airline.xlsx")
export_airline_excel([r3, r2], xlsx)
wb = load_workbook(xlsx); ws = wb.active
assert ws.sheet_view.rightToLeft is False, "يجب أن يكون LTR"
header = [c.value for c in ws[1]]
# الأعمدة الـ11 + عمودا أماديوس
assert header == list(AIRLINE_COLUMNS) + ["Amadeus Name", "Amadeus DOCS"], header
data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]
assert len(data_rows) == 2
# عمود أماديوس DOCS في الصف الأول يطابق التوليد
assert data_rows[0][-1] == amadeus_docs(r3), data_rows[0][-1]
assert data_rows[0][-2] == amadeus_name(r3)
print(f"  OK: إكسل LTR + عمودا أماديوس، {len(data_rows)} صف")

print("\n=== تصدير PDF ===")
pdf = _os.path.join(_OUTDIR, "airline.pdf")
export_airline_pdf([r3, r2], pdf)
assert _os.path.getsize(pdf) > 3000
with open(pdf, "rb") as fh:
    assert fh.read(5) == b"%PDF-"
print(f"  OK: PDF ({_os.path.getsize(pdf)} بايت)")

print("\n=== محدّد الرحلة في كشف الطيران (AirlineDialog) ===")
try:
    from tkinter import Tk
    from hajj_app.gui import AirlineDialog
    root = Tk(); root.withdraw()
    recs = [
        rec(surname_en="ALSHAMSI", given_names_en="ABDULLA", passport_number="A1",
            sex="ذكر", nationality="ARE", airline="الاتحاد", pnr="P1"),
        rec(surname_en="ALNEYADI", given_names_en="MARYAM", passport_number="A2",
            sex="أنثى", nationality="ARE", airline="طيران الإمارات", pnr="P2"),
        rec(surname_en="ALDHAHERI", given_names_en="SALEM", passport_number="A3",
            sex="ذكر", nationality="ARE", airline="الاتحاد", pnr="P3"),
        rec(surname_en="ALKAABI", given_names_en="NORA", passport_number="A4",
            sex="أنثى", nationality="ARE", airline="", pnr="P4"),
    ]
    dlg = AirlineDialog(root, recs)
    # افتراضياً: كل الرحلات
    assert dlg._flight_var.get() == AirlineDialog._ALL_FLIGHTS
    assert len(dlg._current()) == 4, len(dlg._current())
    vals = list(dlg._flight_combo_values())
    # القيم = «كل الرحلات» + الرحلات الفريدة غير الفارغة مرتّبة
    assert vals[0] == AirlineDialog._ALL_FLIGHTS
    assert set(vals[1:]) == {"الاتحاد", "طيران الإمارات"}, vals
    # اختيار رحلة معيّنة -> يصفّي الركّاب والعدّاد والقائمة
    dlg._flight_var.set("الاتحاد"); dlg._rebuild()
    cur = dlg._current()
    assert len(cur) == 2 and all(r.airline == "الاتحاد" for r in cur), cur
    assert len(dlg._tree.get_children()) == 2
    assert "عدد الركّاب: 2" in dlg._count_label.cget("text")
    # اسم الملف الافتراضي يتضمّن اسم الرحلة
    assert "الاتحاد" in dlg._default
    dlg.destroy(); root.destroy()
    print("  OK: الافتراضي كل الرحلات، والتصفية حسب الطيران المختار تعمل")
except Exception as exc:                       # بيئة بلا واجهة رسومية
    if "no display" in str(exc).lower() or "tcl" in type(exc).__name__.lower():
        print(f"  تخطٍّ (لا واجهة رسومية): {exc}")
    else:
        raise

print("\n*** AIRLINE TESTS PASSED ***")

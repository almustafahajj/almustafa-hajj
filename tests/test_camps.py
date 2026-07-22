# -*- coding: utf-8 -*-
"""اختبار كشف تسكين المخيمات: الفصل بالجنس، تماسك العائلة والغرفة، والتصدير."""
import sys, io
import os as _os
_HERE = _os.path.dirname(_os.path.abspath(__file__))
_ROOT = _os.path.dirname(_HERE)
_OUTDIR = _os.path.join(_HERE, "_out")
_os.makedirs(_OUTDIR, exist_ok=True)
sys.path.insert(0, _ROOT)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from hajj_app.camps import (
    CAMP_ARAFAT, CAMP_MINA, MEN, WOMEN, UNKNOWN, CAMP_COLUMNS,
    build_camp_plan, camp_rows, classification, export_camp_excel, tent_label,
)
from hajj_app.pdf_io import export_camp_pdf
from hajj_app.mrz import PassportData


def rec(name, sex, fam="", hotel="", room="", rtype="", phone=""):
    return PassportData(full_name_ar=name, sex=sex, family_number=fam,
                        hotel=hotel, room_number=room, room_type=rtype, phone=phone)


def tents_of(plan, cls):
    return [t for t in plan.tents if t.classification == cls]


def names_in(tent):
    return {o.name for o in tent.occupants}


print("=== التصنيف ===")
assert classification("ذكر") == MEN
assert classification("أنثى") == WOMEN
assert classification("M") == MEN and classification("F") == WOMEN
assert classification("") == UNKNOWN
print("  OK: ذكر→رجال، أنثى→نساء، فراغ→غير محدد")

print("\n=== الفصل بين الرجال والنساء ===")
recs = [
    rec("رجل ١", "ذكر"), rec("امرأة ١", "أنثى"),
    rec("رجل ٢", "ذكر"), rec("امرأة ٢", "أنثى"),
]
plan = build_camp_plan(recs, CAMP_MINA, capacity=10, sector="أ", start_number=1)
for tent in plan.tents:
    sexes = {classification(o.sex) for o in tent.occupants}
    assert len(sexes) == 1, ("خيمة مختلطة!", sexes)
assert {t.classification for t in plan.tents} == {MEN, WOMEN}
assert plan.total == 4
print(f"  OK: {len(plan.tents)} خيمة، لا اختلاط، المخيّم={plan.camp}")

print("\n=== العائلة تبقى معاً في خيمة واحدة ===")
# عائلة 7 من 5 رجال + سعة 4 -> تنقسم لكنها لا تختلط بغيرها إن أمكن
fam = [rec(f"فرد {i}", "ذكر", fam="7") for i in range(3)]
other = [rec("غريب", "ذكر", fam="9")]
plan = build_camp_plan(fam + other, CAMP_MINA, capacity=4, sector="ب")
men_tents = tents_of(plan, MEN)
# الكل رجال ويتّسعون في خيمة واحدة (4) -> خيمة واحدة
assert len(men_tents) == 1 and men_tents[0].count == 4
print("  OK: العائلة والغريب في خيمة واحدة ضمن السعة")

print("\n=== سكّان الغرفة الواحدة معاً ولو اختلفت العائلة ===")
recs = [
    rec("أ", "ذكر", fam="1", hotel="فندق", room="101"),
    rec("ب", "ذكر", fam="2", hotel="فندق", room="101"),   # نفس الغرفة، عائلة أخرى
    rec("ج", "ذكر", fam="3", hotel="فندق", room="202"),
]
plan = build_camp_plan(recs, CAMP_MINA, capacity=2, sector="ج")
# أ و ب (نفس الغرفة) يجب أن يكونا في خيمة واحدة
tent_of = {}
for t in plan.tents:
    for o in t.occupants:
        tent_of[o.name] = id(t)
assert tent_of["أ"] == tent_of["ب"], "سكّان الغرفة تفرّقوا"
assert tent_of["ج"] != tent_of["أ"], "غرفة مختلفة اندمجت خطأً"
print("  OK: سكّان الغرفة 101 معاً، وساكن 202 منفصل")

print("\n=== الفصل بالجنس يتقدّم على العائلة ===")
# عائلة واحدة فيها رجل وامرأة -> خيمتان مختلفتان
recs = [rec("أب", "ذكر", fam="5"), rec("أم", "أنثى", fam="5")]
plan = build_camp_plan(recs, CAMP_MINA, capacity=10)
assert len(plan.tents) == 2
assert names_in(tents_of(plan, MEN)[0]) == {"أب"}
assert names_in(tents_of(plan, WOMEN)[0]) == {"أم"}
print("  OK: العائلة المختلطة انقسمت خيمة رجال وخيمة نساء")

print("\n=== السعة والترقيم والقطاع ===")
recs = [rec(f"رجل {i}", "ذكر") for i in range(5)] + \
       [rec(f"امرأة {i}", "أنثى") for i in range(3)]
plan = build_camp_plan(recs, CAMP_ARAFAT, capacity=2, sector="د", start_number=10)
for t in plan.tents:
    assert t.count <= 2, ("تجاوز السعة", t.count)
    assert t.sector == "د"
    assert t.camp == CAMP_ARAFAT
numbers = [int(t.number) for t in plan.tents]
assert numbers == list(range(10, 10 + len(plan.tents))), numbers   # تسلسلي من 10
# 5 رجال في سعة 2 -> 3 خيام، 3 نساء -> 2 خيمة
assert len(tents_of(plan, MEN)) == 3 and len(tents_of(plan, WOMEN)) == 2
print(f"  OK: خيام مرقّمة {numbers}, قطاع د, مخيّم عرفة, بلا تجاوز سعة")

print("\n=== غير محدّد الجنس: خيمة منفصلة + تنبيه ===")
recs = [rec("مجهول", ""), rec("رجل", "ذكر")]
plan = build_camp_plan(recs, CAMP_MINA, capacity=10)
assert len(tents_of(plan, UNKNOWN)) == 1
assert any("غير محدد" in n or "جنس" in n for n in plan.notes), plan.notes
print(f"  OK: خيمة «غير محدد» + تنبيه: {plan.notes}")

print("\n=== وحدة أكبر من سعة الخيمة تُقسَّم مع تنبيه ===")
recs = [rec(f"فرد {i}", "ذكر", fam="1") for i in range(5)]   # عائلة 5 في سعة 2
plan = build_camp_plan(recs, CAMP_MINA, capacity=2)
assert len(tents_of(plan, MEN)) == 3   # 2+2+1
assert any("تتجاوز سعة" in n for n in plan.notes), plan.notes
print(f"  OK: عائلة 5 في سعة 2 -> 3 خيام + تنبيه")

print("\n=== اختيار الجنس (only) ===")
recs = [rec(f"رجل {i}", "ذكر") for i in range(3)] + \
       [rec(f"امرأة {i}", "أنثى") for i in range(2)]
men_only = build_camp_plan(recs, CAMP_MINA, capacity=2, sector="أ", start_number=1, only=MEN)
assert {t.classification for t in men_only.tents} == {MEN}
assert men_only.total == 3
# خيام الرجال وحدها تُرقّم من start_number (1) لا بعد النساء
assert [int(t.number) for t in men_only.tents] == [1, 2]
women_only = build_camp_plan(recs, CAMP_MINA, capacity=5, only=WOMEN)
assert {t.classification for t in women_only.tents} == {WOMEN}
assert women_only.total == 2
# تصنيف غير معروف يعني الكل
both = build_camp_plan(recs, CAMP_MINA, capacity=5, only="xyz")
assert {t.classification for t in both.tents} == {MEN, WOMEN}
# only=نساء يُخفي تنبيه غير المحدّد
mix = [rec("مجهول", ""), rec("امرأة", "أنثى")]
w = build_camp_plan(mix, CAMP_MINA, only=WOMEN)
assert all("غير محدد" not in n for n in w.notes), w.notes
print("  OK: رجال وحدهم/نساء وحدهن، ترقيم من البداية، والكل عند تصنيف مجهول")

print("\n=== صفوف العرض ===")
recs = [rec("أحمد", "ذكر", fam="1", hotel="الصفا", room="12", phone="0555")]
plan = build_camp_plan(recs, CAMP_MINA, capacity=4, sector="أ")
rows = camp_rows(plan)
assert len(rows) == 1
r = rows[0]
assert r["name"] == "أحمد" and r["sector"] == "أ" and r["classification"] == MEN
assert r["hotel"] == "الصفا" and r["room"] == "12" and r["phone"] == "0555"
assert "خيمة" in tent_label(plan.tents[0]) and "رجال" in tent_label(plan.tents[0])
print(f"  OK: {tent_label(plan.tents[0])}")

print("\n=== تصدير إكسل ===")
recs = [rec(f"رجل {i}", "ذكر", fam="1") for i in range(3)] + \
       [rec(f"امرأة {i}", "أنثى", fam="2") for i in range(2)]
plan = build_camp_plan(recs, CAMP_MINA, capacity=4, sector="أ", start_number=1)
xlsx = _os.path.join(_OUTDIR, "camps.xlsx")
export_camp_excel(plan, xlsx)
wb = load_workbook(xlsx)
ws = wb.active
assert ws.sheet_view.rightToLeft is True
assert ws.title == CAMP_MINA
header = [c.value for c in ws[2]]
assert header == list(CAMP_COLUMNS), header
# عدد صفوف الحجّاج = المجموع (5)، وصفوف عناوين الخيام = عدد الخيام
values = [row[0].value for row in ws.iter_rows(min_row=3)]
serials = [v for v in values if isinstance(v, int)]
assert len(serials) == 5, serials
print(f"  OK: إكسل RTL، ورقة={ws.title}، {len(serials)} حاجاً + عناوين خيام")

print("\n=== تصدير PDF ===")
pdf = _os.path.join(_OUTDIR, "camps.pdf")
export_camp_pdf(plan, pdf)
assert _os.path.getsize(pdf) > 3000
with open(pdf, "rb") as fh:
    assert fh.read(5) == b"%PDF-"
print(f"  OK: PDF ({_os.path.getsize(pdf)} بايت)")

print("\n=== نافذة المخيمات (CampsDialog) ===")
try:
    from tkinter import Tk
    from hajj_app.gui import CampsDialog
    root = Tk(); root.withdraw()
    recs = [rec("رجل ١", "ذكر", fam="1"), rec("امرأة ١", "أنثى", fam="2"),
            rec("رجل ٢", "ذكر", fam="1")]
    dlg = CampsDialog(root, recs)
    dlg._cap_var.set("2"); dlg._sector_var.set("ب"); dlg._camp_var.set(CAMP_ARAFAT)
    dlg._rebuild(force=True)
    assert dlg._plan.camp == CAMP_ARAFAT and dlg._plan.capacity == 2
    assert dlg._plan.total == 3
    # لا اختلاط
    for t in dlg._plan.tents:
        assert len({classification(o.sex) for o in t.occupants}) == 1
    # تعديل يدوي لرقم خيمة يبقى ولا يُمحى ما لم تتغيّر المعطيات
    first = dlg._plan.tents[0]
    first.number = "99"; dlg._refresh_labels()
    dlg._rebuild(force=False)   # نفس المعطيات -> لا إعادة بناء
    assert dlg._plan.tents[0].number == "99", "التعديل اليدوي مُحي بلا تغيير معطيات"
    # تغيير المعطيات يعيد البناء (يُلغي التجاوز اليدوي)
    dlg._start_var.set("5"); dlg._rebuild(force=False)
    assert dlg._plan.tents[0].number == "5"
    # اختيار الجنس: نساء فقط
    dlg._class_var.set(WOMEN); dlg._rebuild(force=False)
    assert {t.classification for t in dlg._plan.tents} == {WOMEN}
    assert dlg._plan.total == 1
    dlg._class_var.set(dlg._ALL_CLASSES); dlg._rebuild(force=False)
    assert dlg._plan.total == 3
    dlg.destroy(); root.destroy()
    print("  OK: التوزيع، الفصل، اختيار الجنس، وثبات التعديل اليدوي")
except Exception as exc:
    if "no display" in str(exc).lower() or "tcl" in type(exc).__name__.lower():
        print(f"  تخطٍّ (لا واجهة رسومية): {exc}")
    else:
        raise

print("\n*** CAMPS TESTS PASSED ***")

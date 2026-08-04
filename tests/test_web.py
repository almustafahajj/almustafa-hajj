# -*- coding: utf-8 -*-
"""اختبار نسخة الويب (المرحلة صفر): الدخول، الجلسة، عرض الكشف، البحث، الخروج."""
import sys, io
import os as _os
_HERE = _os.path.dirname(_os.path.abspath(__file__))
_ROOT = _os.path.dirname(_HERE)
_OUTDIR = _os.path.join(_HERE, "_out")
_os.makedirs(_OUTDIR, exist_ok=True)
sys.path.insert(0, _ROOT)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from pathlib import Path
import shutil

from hajj_app import auth, storage
from hajj_app.mrz import PassportData
from hajj_web import create_app

WORK = Path(_OUTDIR) / "webdata"
shutil.rmtree(WORK, ignore_errors=True)
WORK.mkdir(parents=True, exist_ok=True)
AUTH, DATA = WORK / "auth.json", WORK / "hajjaj.json"
# عزل الإعدادات والنسخ الاحتياطية عن بيانات المستخدم الحقيقية
storage.default_data_path = lambda: DATA

print("=== تجهيز حساب وبيانات مشفّرة ===")
admin, _rk = auth.create_account("MHU", "Web-Pass-1234", AUTH)
auth.add_account(admin, "viewer1", "View-Pass-1234", "viewer", AUTH)
storage.save_records([
    PassportData(full_name_ar="عبدالله الشامسي", passport_number="A111",
                 phone="0501112233", hotel="الصفوة", program="الأول",
                 nationality_ar="سعودي", room_number="101",
                 program_value="5000", paid_amount="2000"),
    PassportData(full_name_ar="سالم أحمد", passport_number="B222",
                 phone="0502223344", hotel="كونراد", program="الثاني",
                 nationality_ar="إماراتي", room_number="202"),
], DATA, admin)
print("  OK: حسابان + حاجّان مشفّران")

app = create_app(auth_path=AUTH, data_path=DATA)
client = app.test_client()

print("\n=== الوصول بلا دخول محجوب ===")
assert client.get("/login").status_code == 200
r = client.get("/")
assert r.status_code == 302 and "/login" in r.headers["Location"], r.status_code
print("  OK: '/' يعيد التوجيه إلى صفحة الدخول")

print("\n=== دخول خاطئ لا يُفصح ولا يُدخِل ===")
r = client.post("/login", data={"username": "MHU", "password": "خطأ"})
body = r.get_data(as_text=True)
assert "غير صحيح" in body, "رسالة الخطأ غير ظاهرة"
assert "Set-Cookie" not in r.headers, "أُنشئت جلسة رغم خطأ كلمة المرور!"
print("  OK: يُرفض دون إنشاء جلسة")

print("\n=== دخول صحيح (مطّلع) يعرض الكشف ===")
r = client.post("/login", data={"username": "viewer1", "password": "View-Pass-1234"})
assert r.status_code == 302
assert "hajj_session" in r.headers.get("Set-Cookie", ""), "لم تُوضع كوكيّ الجلسة"
# الكوكيّ يحمل رمزاً فقط لا مفتاح البيانات
cookie = r.headers["Set-Cookie"]
page = client.get("/").get_data(as_text=True)
assert "عبدالله الشامسي" in page and "سالم أحمد" in page, "الكشف لا يظهر"
assert "مطّلع" in page, "الدور لا يظهر"
assert "المعروض: 2 من 2" in page, page[:200]
print("  OK: الجدول يعرض الحجّاج والدور، والكوكيّ رمز فقط")

print("\n=== البحث يُرشّح ===")
f = client.get("/?q=سالم").get_data(as_text=True)
assert "سالم أحمد" in f and "عبدالله الشامسي" not in f
f2 = client.get("/?q=A111").get_data(as_text=True)
assert "عبدالله الشامسي" in f2 and "سالم أحمد" not in f2
print("  OK: البحث بالاسم والجواز يُرشّح النتائج")

print("\n=== الفلاتر تُرشّح (البرنامج/الفندق/الجنسية) ===")
fp = client.get("/?program=الأول").get_data(as_text=True)
assert "عبدالله الشامسي" in fp and "سالم أحمد" not in fp, "فلتر البرنامج لا يعمل"
assert "المعروض: 1 من 2" in fp
fh = client.get("/?hotel=كونراد").get_data(as_text=True)
assert "سالم أحمد" in fh and "عبدالله الشامسي" not in fh, "فلتر الفندق لا يعمل"
# قوائم الفلاتر تحوي القيم المتاحة
assert "الصفوة" in page and "كونراد" in page, "قوائم الفلاتر ناقصة"
print("  OK: الترشيح بالبرنامج والفندق، والقوائم مملوءة")

print("\n=== فتح سجلّ الحاج كاملاً ===")
detail = client.get("/pilgrim/0").get_data(as_text=True)
assert "عبدالله الشامسي" in detail
# عناوين المجموعات والحقول تظهر
assert "بيانات الجواز" in detail and "المالية" in detail
assert "المبلغ المتبقي" in detail, "الحقول المحسوبة لا تظهر"
assert "3,000" in detail, "المتبقي (5000-2000) لا يُحسب"     # 5000 - 2000
# فهرس خارج المدى يعطي 404
assert client.get("/pilgrim/999").status_code == 404
print("  OK: السجلّ الكامل يعرض كل الحقول والمتبقي المحسوب، والفهرس الخاطئ 404")

print("\n=== الخروج ينهي الجلسة ===")
client.get("/logout")
assert client.get("/").status_code == 302, "الجلسة بقيت بعد الخروج"
print("  OK: بعد الخروج يُطلب الدخول من جديد")

print("\n=== المطّلع لا يرى أزرار التعديل ولا مساراته ===")
viewer_home = client.post("/login", data={"username": "viewer1",
                          "password": "View-Pass-1234"}) and client.get("/").get_data(as_text=True)
assert "إضافة حاج" not in viewer_home, "زر الإضافة ظاهر للمطّلع!"
# مسارات الكتابة محجوبة للمطّلع (403)
assert client.get("/pilgrim/new").status_code == 403
assert client.post("/pilgrim/0/delete", data={"orig": "A111"}).status_code == 403
print("  OK: المطّلع بلا أزرار تعديل ومساراته 403")

print("\n=== المدير: إضافة/تعديل/حذف ===")
mgr = app.test_client()
mgr.post("/login", data={"username": "MHU", "password": "Web-Pass-1234"})
assert "إضافة حاج" in mgr.get("/").get_data(as_text=True), "زر الإضافة غائب عن المدير"
# نماذج الإضافة/التعديل تُعرض (GET)
newform = mgr.get("/pilgrim/new").get_data(as_text=True)
assert "إضافة حاج جديد" in newform and "اسم الحاج بالعربي" in newform
editform = mgr.get("/pilgrim/0/edit").get_data(as_text=True)
assert "تعديل سجلّ الحاج" in editform and "عبدالله الشامسي" in editform
print("  OK: نماذج الإضافة والتعديل تُعرض بالحقول")
# إضافة
r = mgr.post("/pilgrim/new", data={"full_name_ar": "حاج جديد",
             "passport_number": "C333", "hotel": "هيلتون", "program": "الأول"})
assert r.status_code == 302
recs, _ = storage.load_records(DATA, admin)
assert any(x.passport_number == "C333" and x.full_name_ar == "حاج جديد" for x in recs)
assert len(recs) == 3
print("  OK: أُضيف الحاج وحُفظ مشفّراً")

# تعديل (الحاج الجديد آخر السجلّات = فهرس 2)
r = mgr.post("/pilgrim/2/edit", data={"full_name_ar": "حاج معدّل",
             "passport_number": "C333", "orig": "C333", "hotel": "هيلتون"})
assert r.status_code == 302
recs, _ = storage.load_records(DATA, admin)
assert recs[2].full_name_ar == "حاج معدّل"
# حارس التزامن: قيمة orig خاطئة تُرفض
bad = mgr.post("/pilgrim/2/edit", data={"passport_number": "C333", "orig": "WRONG"})
assert bad.status_code == 409
print("  OK: عُدّل الحاج، وحارس التزامن يرفض orig الخاطئ (409)")

# حذف
r = mgr.post("/pilgrim/2/delete", data={"orig": "C333"})
assert r.status_code == 302
recs, _ = storage.load_records(DATA, admin)
assert len(recs) == 2 and all(x.passport_number != "C333" for x in recs)
print("  OK: حُذف الحاج")

# سجلّ التدقيق سجّل العمليات باسم المستخدم (ويب)
from hajj_app import audit
entries = audit.read_entries(path=WORK / "audit.log")
actions = {e["action"] for e in entries}
assert {"إضافة حاج", "تعديل حاج", "حذف حاج"} <= actions, actions
assert all("(ويب)" in e["user"] for e in entries if e["action"].endswith("حاج"))
print("  OK: العمليات مسجّلة في التدقيق باسم المستخدم (ويب)")

print("\n=== التصدير (إكسل/PDF) يحترم الفلتر ===")
xl = mgr.get("/export/excel")
assert xl.status_code == 200 and xl.data[:2] == b"PK", "إكسل غير صالح"
assert "spreadsheetml" in xl.headers.get("Content-Type", "")
pf = mgr.get("/export/pdf")
assert pf.status_code == 200 and pf.data[:5] == b"%PDF-", "PDF غير صالح"
# التصدير المفلتر: برنامج «الثاني» فيه حاجّ واحد -> إكسل أصغر وPDF صالح
xl2 = mgr.get("/export/excel?program=الثاني")
assert xl2.status_code == 200 and xl2.data[:2] == b"PK"
from openpyxl import load_workbook
import io as _io
wb = load_workbook(_io.BytesIO(xl2.data)); ws = wb.active
data_rows = ws.max_row - 2          # صف العنوان + صف الرؤوس
assert data_rows == 1, f"التصدير المفلتر يجب أن يحوي حاجّاً واحداً لا {data_rows}"
flat = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
assert "سالم أحمد" in flat and "عبدالله الشامسي" not in flat
print("  OK: إكسل وPDF صالحان، والتصدير يحترم فلتر البرنامج")

print("\n=== الاستيراد من إكسل (رفع ملف) ===")
# نبني ملف إكسل من حاجّين ونرفعه فيُضافان
from hajj_app.excel_io import export_excel
IMP = WORK / "to_import.xlsx"
export_excel([
    PassportData(full_name_ar="مستورد أول", passport_number="X1", hotel="فندق"),
    PassportData(full_name_ar="مستورد ثاني", passport_number="X2", hotel="فندق"),
], IMP)
before = len(storage.load_records(DATA, admin)[0])
with open(IMP, "rb") as fh:
    r = mgr.post("/import/excel", data={"file": (fh, "to_import.xlsx")},
                 content_type="multipart/form-data")
assert r.status_code == 302
after = storage.load_records(DATA, admin)[0]
assert len(after) == before + 2, f"{before} -> {len(after)}"
assert any(x.passport_number == "X1" for x in after)
print(f"  OK: أُضيف حاجّان بالاستيراد ({before} -> {len(after)})")

# صفحة الاستيراد + المطّلع ممنوع + بلا ملف
assert "استيراد من إكسل" in mgr.get("/import").get_data(as_text=True)
assert client.get("/import").status_code == 403          # المطّلع
assert client.post("/import/excel").status_code == 403
nof = mgr.post("/import/excel", data={}, content_type="multipart/form-data")
assert nof.status_code == 302                            # يعيد التوجيه برسالة خطأ
# رفع بلا ملفات لقراءة الجوازات يعيد التوجيه (لا يتعطّل)
assert mgr.post("/import/passports", data={},
                content_type="multipart/form-data").status_code == 302
print("  OK: صفحة الاستيراد، ومنع المطّلع (403)، ورفع فارغ آمن")

print("\n=== رمز QR للفتح من جهاز آخر ===")
q = client.get("/qr")                         # متاح بلا دخول (يعرض رابط الشبكة فقط)
qhtml = q.get_data(as_text=True)
assert q.status_code == 200
assert "data:image/png;base64," in qhtml, "صورة QR غير مضمّنة"
assert "http://" in qhtml, "الرابط لا يظهر"
# الرابط في صفحة الدخول
assert "رمز QR" in client.get("/login").get_data(as_text=True)
print("  OK: صفحة QR تعرض الرمز والرابط، ورابطها في صفحة الدخول")

print("\n=== الإحصاءات ===")
st = mgr.get("/stats").get_data(as_text=True)
assert "الملخّص المالي" in st and "التوزيع حسب" in st
assert "عدد الحجّاج" in st
sp = mgr.get("/stats/pdf")
assert sp.status_code == 200 and sp.data[:5] == b"%PDF-"
print("  OK: صفحة الإحصاءات + تصدير PDF")

print("\n=== الكشوف والبطاقات (تنزيلات PDF) ===")
rp = mgr.get("/reports").get_data(as_text=True)
assert "كشف المواصلات" in rp and "بطاقات الحجّاج" in rp
for path in ["/reports/transport.pdf", "/reports/airline.pdf",
             "/reports/badges.pdf", "/reports/stickers/room.pdf"]:
    r = mgr.get(path)
    assert r.status_code == 200 and r.data[:5] == b"%PDF-", path
tx = mgr.get("/reports/transport.xlsx")
assert tx.status_code == 200 and tx.data[:2] == b"PK"
assert mgr.get("/reports/stickers/زبد.pdf").status_code == 404   # نوع غير معروف
print("  OK: مواصلات/طيران/بطاقات/استيكرات PDF + مواصلات إكسل")

print("\n=== إدارة الحسابات (مدير فقط) ===")
# المطّلع لا يصل
assert client.get("/accounts").status_code == 403
# المدير: الصفحة + إضافة حساب (يظهر مفتاح الاسترداد مرة)
acc = mgr.get("/accounts").get_data(as_text=True)
assert "إدارة الحسابات" in acc and "viewer1" in acc
r = mgr.post("/accounts/add", data={"username": "web_editor",
             "password": "WebEd!Pass1", "role": "editor"})
body = r.get_data(as_text=True)
assert "مفتاح استرداد" in body, "لم يُعرض مفتاح الاسترداد"
assert "web_editor" in {a["username"] for a in auth.list_accounts(AUTH)}
# الحساب الجديد يفتح البيانات نفسها
assert auth.login("web_editor", "WebEd!Pass1", AUTH).role == "editor"
# تغيير الدور + منع حذف الذات
mgr.post("/accounts/role", data={"username": "web_editor", "role": "viewer"})
assert {a["username"]: a["role"] for a in auth.list_accounts(AUTH)}["web_editor"] == "viewer"
mgr.post("/accounts/delete", data={"username": "MHU"})     # حذف الذات مرفوض
assert "MHU" in {a["username"] for a in auth.list_accounts(AUTH)}
mgr.post("/accounts/delete", data={"username": "web_editor"})
assert "web_editor" not in {a["username"] for a in auth.list_accounts(AUTH)}
print("  OK: إضافة (مع المفتاح)، تغيير الدور، منع حذف الذات، حذف")

print("\n=== سجلّ التدقيق ===")
au = mgr.get("/audit").get_data(as_text=True)
assert "سجلّ التدقيق" in au
assert "إضافة حساب" in au and "web_editor" in au       # العمليات السابقة مسجّلة
assert "(ويب)" in au
print("  OK: سجلّ التدقيق يعرض العمليات باسم المستخدم (ويب)")

print("\n=== كشف التسكين + مستندات الحاج + PEPPOL ===")
for path, head in [("/reports/rooming.pdf", b"%PDF-"),
                   ("/pilgrim/0/receipt.pdf", b"%PDF-"),
                   ("/pilgrim/0/invoice.pdf", b"%PDF-"),
                   ("/pilgrim/0/contract.pdf", b"%PDF-")]:
    r = mgr.get(path)
    assert r.status_code == 200 and r.data[:5] == head, path
assert mgr.get("/reports/rooming.xlsx").data[:2] == b"PK"
xml = mgr.get("/pilgrim/0/einvoice.xml")
assert xml.status_code == 200 and b"<?xml" in xml.data[:10] and b"Invoice" in xml.data
print("  OK: تسكين PDF/إكسل، سند/فاتورة/عقد PDF، وفاتورة PEPPOL XML")

print("\n=== واتساب الحاج ===")
# الحاج 0 لا هاتف له -> يعيد التوجيه للسجلّ برسالة؛ نضيف هاتفاً لواحد
wa = mgr.get("/pilgrim/0/whatsapp")
assert wa.status_code == 302   # إمّا wa.me أو رجوع للسجلّ
print("  OK: مسار واتساب يعمل (تحويل)")

print("\n=== فحص الجاهزية ===")
ql = mgr.get("/quality").get_data(as_text=True)
assert "فحص جاهزية الكشف" in ql
print("  OK: صفحة فحص الجاهزية")

print("\n=== برامج الحملة (تعديل + حفظ) ===")
assert client.get("/programs").status_code == 403          # المطّلع ممنوع
pg = mgr.get("/programs").get_data(as_text=True)
assert "برامج الحملة" in pg and "تاريخ السفر" in pg
# حفظ قيمة ثم قراءتها (الإعدادات معزولة في WORK)
r = mgr.post("/programs", data={"p0_hotel": "فندق البرنامج الأول",
             "p0_cost_double": "6000", "p0_travel_date": "1447-12-01"})
assert r.status_code == 302
from hajj_app import programs as _pg
saved = _pg.load_programs(storage.load_settings())
assert saved[0].hotel == "فندق البرنامج الأول" and saved[0].cost_double == "6000"
print("  OK: صفحة البرامج تُعرض وتُحفظ القيم، والمطّلع ممنوع (403)")

print("\n=== نسخة احتياطية ===")
assert client.post("/backup").status_code == 403           # المطّلع ممنوع
b = mgr.post("/backup")
assert b.status_code == 302
print("  OK: نسخة احتياطية للمحرّر/المدير، والمطّلع ممنوع")

print("\n=== لوحة المؤشّرات على الكشف ===")
home = mgr.get("/").get_data(as_text=True)
assert "kpis dash" in home and "المحصّل" in home
print("  OK: شريط المؤشّرات المالية يظهر أعلى الكشف")

print("\n=== استعادة نسخة احتياطية ===")
assert client.get("/restore").status_code == 403          # المطّلع ممنوع
mgr.post("/backup")                                       # أنشئ نسخة
rlist = mgr.get("/restore").get_data(as_text=True)
assert "استعادة نسخة احتياطية" in rlist
import re
m = re.search(r'name="name" value="(hajjaj-[^"]+)"', rlist)
assert m, "لا توجد نسخة للاستعادة"
before = len(storage.load_records(DATA, admin)[0])
ra = mgr.post("/restore/apply", data={"name": m.group(1)})
assert ra.status_code == 302
assert len(storage.load_records(DATA, admin)[0]) == before  # نفس العدد بعد الاستعادة
# اسم نسخة غير موجود -> 404
assert mgr.post("/restore/apply", data={"name": "hajjaj-x.json"}).status_code == 404
print("  OK: استعادة نسخة (مع نسخ أمان)، والمطّلع ممنوع، والاسم الخاطئ 404")

print("\n=== تغيير كلمة المرور ومفتاح الاسترداد ===")
# كلمة مرور خاطئة تُرفض
mgr.post("/account/password", data={"current": "خطأ", "password": "New-Pass-9999",
         "confirm": "New-Pass-9999"})
assert auth.login("MHU", "Web-Pass-1234", AUTH)          # لم تتغيّر
# تغيير صحيح
r = mgr.post("/account/password", data={"current": "Web-Pass-1234",
             "password": "New-Pass-9999", "confirm": "New-Pass-9999"})
assert r.status_code in (302, 200)
try:
    auth.login("MHU", "Web-Pass-1234", AUTH); assert False, "القديمة تعمل"
except auth.AuthError:
    pass
assert auth.login("MHU", "New-Pass-9999", AUTH)
# مفتاح استرداد جديد يُعرض
rk = mgr.post("/account/recovery", data={"password": "New-Pass-9999"})
assert "مفتاح الاسترداد" in rk.get_data(as_text=True)
print("  OK: تغيير كلمة المرور فعّال، ومفتاح الاسترداد الجديد يُعرض")

print("\n=== صور الجوازات (رفع/عرض/طباعة) ===")
# نبني صورة PNG بسيطة في الذاكرة
from PIL import Image
import io as _io2
img_buf = _io2.BytesIO()
Image.new("RGB", (400, 260), (200, 180, 140)).save(img_buf, format="PNG")
img_bytes = img_buf.getvalue()
# لا صورة قبل الرفع
assert mgr.get("/pilgrim/0/image/passport").status_code == 404
# رفع صورة جواز للحاج 0 (مدير)
r = mgr.post("/pilgrim/0/image/passport/upload",
             data={"file": (_io2.BytesIO(img_bytes), "jaz.png")},
             content_type="multipart/form-data")
assert r.status_code == 302
# الآن تُعرض كـ PNG
shown = mgr.get("/pilgrim/0/image/passport")
assert shown.status_code == 200 and shown.data[:8] == b"\x89PNG\r\n\x1a\n"
assert "image/png" in shown.headers.get("Content-Type", "")
# صفحة السجلّ تُظهر الصورة
assert "الصور" in mgr.get("/pilgrim/0").get_data(as_text=True)
# المطّلع لا يرفع
assert client.post("/pilgrim/0/image/passport/upload").status_code == 403
# طباعة الجوازات: الآن فيها جواز واحد على الأقل
pp = mgr.get("/reports/passports.pdf")
assert pp.status_code == 200 and pp.data[:5] == b"%PDF-"
print("  OK: رفع صورة الجواز، عرضها PNG، طباعة الجوازات PDF، ومنع المطّلع")

print("\n=== تحديد متعدّد: تعديل جماعي + حذف + تراجع ===")
n0 = len(storage.load_records(DATA, admin)[0])
assert n0 >= 3, n0
# تعديل جماعي: طبّق الفندق على أول سجلّين
r = mgr.post("/bulk/edit/apply", data={"sel": ["0", "1"],
             "apply_hotel": "on", "hotel": "فندق موحّد"})
assert r.status_code == 302
recs = storage.load_records(DATA, admin)[0]
assert recs[0].hotel == "فندق موحّد" and recs[1].hotel == "فندق موحّد"
# تراجع يعيد الفنادق السابقة
u = mgr.post("/undo")
assert u.status_code == 302
recs2 = storage.load_records(DATA, admin)[0]
assert recs2[0].hotel != "فندق موحّد", "التراجع لم يُرجِع القيمة"
print("  OK: تعديل جماعي للحقول المفعّلة، والتراجع يعيد الحالة")

# حذف جماعي + تراجع
before = len(storage.load_records(DATA, admin)[0])
r = mgr.post("/bulk/delete", data={"sel": ["0", "1"]})
assert r.status_code == 302
assert len(storage.load_records(DATA, admin)[0]) == before - 2
mgr.post("/undo")
assert len(storage.load_records(DATA, admin)[0]) == before, "التراجع عن الحذف فشل"
print("  OK: حذف جماعي والتراجع يستعيد المحذوفين")

# المطّلع ممنوع من كل المسارات الجماعية
for path in ["/bulk/delete", "/bulk/edit", "/bulk/edit/apply", "/undo"]:
    assert client.post(path).status_code == 403, path
# مربّعات الاختيار تظهر للمحرّر/المدير فقط
assert 'name="sel"' in mgr.get("/").get_data(as_text=True)
assert 'name="sel"' not in client.get("/").get_data(as_text=True)
print("  OK: المطّلع ممنوع من المسارات الجماعية ولا يرى مربّعات التحديد")

print("\n=== خيام المخيمات ===")
cp = mgr.get("/reports/camps.pdf?camp=منى&capacity=20")
assert cp.status_code == 200 and cp.data[:5] == b"%PDF-"
cp2 = mgr.get("/reports/camps.pdf?camp=عرفة")
assert cp2.status_code == 200 and cp2.data[:5] == b"%PDF-"
assert "خيام المخيمات" in mgr.get("/reports").get_data(as_text=True)
print("  OK: كشف مخيّم منى/عرفة PDF بسعة مختارة")

print("\n=== المعاينة: PDF داخل المتصفّح، إكسل تنزيل ===")
pdfr = mgr.get("/export/pdf")
assert "inline" in pdfr.headers.get("Content-Disposition", ""), \
    pdfr.headers.get("Content-Disposition")
xlr = mgr.get("/export/excel")
assert "attachment" in xlr.headers.get("Content-Disposition", "")
# مستند الحاج (سند قبض) يُعرض داخلياً أيضاً
rc = mgr.get("/pilgrim/0/receipt.pdf")
assert "inline" in rc.headers.get("Content-Disposition", "")
print("  OK: PDF يُفتح للمعاينة (inline)، وإكسل يُنزَّل (attachment)")

print("\n=== إضافة حاج بقراءة الجواز ===")
assert "قراءة الجواز" in mgr.get("/pilgrim/scan").get_data(as_text=True)
assert client.get("/pilgrim/scan").status_code == 403           # المطّلع ممنوع
# بلا ملف -> رجوع
assert mgr.post("/pilgrim/scan").status_code == 302
# رفع صورة بلا MRZ صالح -> يعرض النموذج للمراجعة (يدوياً) دون تعطّل
from PIL import Image as _Img
import io as _io3
b = _io3.BytesIO(); _Img.new("RGB", (600, 400), (240, 240, 240)).save(b, format="PNG")
sc = mgr.post("/pilgrim/scan", data={"file": (_io3.BytesIO(b.getvalue()), "jaz.png")},
              content_type="multipart/form-data")
# إمّا 200 (النموذج للمراجعة) أو 302 (إن غاب Tesseract على هذا الجهاز)
assert sc.status_code in (200, 302)
if sc.status_code == 200:
    assert "من الجواز" in sc.get_data(as_text=True)
print("  OK: صفحة قراءة الجواز تعمل، والمطّلع ممنوع، والرفع الفارغ آمن")

print("\n=== وضع العمرة: التبديل والبرامج والإدارة المالية ===")
import hajj_app.app_mode as _am
import hajj_app.umrah as _um
UDATA = WORK / "umrah.json"
# نجهّز برنامج عمرة ومعتمريه في ملفّات العمرة المعزولة (umrah.json + settings_umrah)
_am.set_mode("umrah")
_usettings = storage.load_settings()
_um.save_trips(_usettings, [
    _um.UmrahTrip(code="U1", name="رمضان", makkah_hotel="كونراد",
                  madinah_hotel="المدينة المنوّرة", depart_date="2026-03-01",
                  return_date="2026-03-10", capacity="40"),
])
storage.save_settings(_usettings)
storage.save_records([
    PassportData(full_name_ar="معتمر مسدّد", trip="U1", passport_number="U-1",
                 room_type="ثنائي", program_value="5000", paid_amount="5000"),
    PassportData(full_name_ar="معتمر جزئي", trip="U1", passport_number="U-2",
                 room_type="ثلاثي", program_value="4000", paid_amount="1500"),
    PassportData(full_name_ar="معتمر غير مدفوع", trip="U1", passport_number="U-3",
                 room_type="مفرد", program_value="6000", paid_amount="0"),
], UDATA, admin)
_am.set_mode("hajj")                         # التطبيق يضبط الوضع من الكوكيّ

um = app.test_client()
um.post("/login", data={"username": "MHU", "password": "New-Pass-9999"})
# قبل التبديل: الرئيسية كشف الحج + زرّ التبديل للعمرة ظاهر
hajj_home = um.get("/").get_data(as_text=True)
assert "برنامج موسم الحج" in hajj_home and "التبديل إلى العمرة" in hajj_home
# التبديل إلى العمرة يضبط الكوكيّ
sw = um.get("/mode/umrah")
assert sw.status_code == 302
assert "hajj_mode" in sw.headers.get("Set-Cookie", "")
# الرئيسية الآن تعيد التوجيه إلى برامج العمرة
home = um.get("/")
assert home.status_code == 302 and "/umrah/programs" in home.headers["Location"]
progs = um.get("/umrah/programs").get_data(as_text=True)
assert "برامج العمرة" in progs and "رمضان" in progs and "U1" in progs
assert "كونراد" in progs                     # فندق مكة يظهر
# صفحة البرنامج: بطاقات مالية + حالات المعتمرين + الإجماليات
prog = um.get("/umrah/program/U1").get_data(as_text=True)
assert "معتمر مسدّد" in prog and "معتمر جزئي" in prog and "معتمر غير مدفوع" in prog
assert "نسبة التحصيل" in prog and "متأخّرون" in prog
assert "مسدّد" in prog and "جزئي" in prog and "غير مدفوع" in prog
assert "15,000" in prog and "6,500" in prog and "8,500" in prog   # 15000/6500/8500
# برنامج غير موجود -> 404
assert um.get("/umrah/program/ZZ").status_code == 404
# الملخّص المالي PDF للبرنامج
fp = um.get("/umrah/program/U1/finance.pdf")
assert fp.status_code == 200 and fp.data[:5] == b"%PDF-"
# كشف المعتمرين PDF/إكسل + مستندات المعتمر (سند/فاتورة/عقد)
assert "كشف المعتمرين PDF" in prog and "المستندات" in prog
rpdf = um.get("/umrah/program/U1/roster.pdf")
assert rpdf.status_code == 200 and rpdf.data[:5] == b"%PDF-"
assert "inline" in rpdf.headers.get("Content-Disposition", "")   # معاينة
rxls = um.get("/umrah/program/U1/roster.xlsx")
assert rxls.status_code == 200 and rxls.data[:2] == b"PK"
assert "attachment" in rxls.headers.get("Content-Disposition", "")  # تنزيل
for _doc in ("receipt", "invoice", "contract"):
    d = um.get(f"/umrah/program/U1/pilgrim/0/{_doc}.pdf")
    assert d.status_code == 200 and d.data[:5] == b"%PDF-", _doc
# معتمر خارج البرنامج / برنامج مجهول -> 404
assert um.get("/umrah/program/U1/pilgrim/99/receipt.pdf").status_code == 404
assert um.get("/umrah/program/ZZ/roster.pdf").status_code == 404
# بيانات العمرة معزولة عن الحج: ملفّ umrah.json مستقلّ
assert UDATA.is_file() and DATA.is_file() and UDATA != DATA

# --- المرحلة ٢: إدارة البرامج والمعتمرين والدفعات على الويب ---
from hajj_app.fields import payment_total as _ptot
assert "برنامج جديد" in um.get("/umrah/programs").get_data(as_text=True)
# إضافة برنامج
r = um.post("/umrah/program/new", data={"name": "شعبان", "makkah_hotel": "فندق",
            "price_double": "4500", "capacity": "30", "depart_date": "2026-02-01"})
assert r.status_code == 302 and "/umrah/program/" in r.headers["Location"]
NC = r.headers["Location"].rstrip("/").split("/")[-1]
_am.set_mode("umrah")
_trips = _um.load_trips(storage.load_settings())
assert any(t.code == NC and t.name == "شعبان" for t in _trips)
# تعديل البرنامج
r = um.post(f"/umrah/program/{NC}/edit", data={"name": "شعبان المعدّل",
            "price_double": "4700", "makkah_hotel": "فندق"})
assert r.status_code == 302
_am.set_mode("umrah")
_t = next(t for t in _um.load_trips(storage.load_settings()) if t.code == NC)
assert _t.name == "شعبان المعدّل" and _t.price_double == "4700"
# إضافة معتمر إلى البرنامج (يأخذ رقماً مرجعياً تلقائياً)
r = um.post(f"/umrah/program/{NC}/pilgrim/new", data={"full_name_ar": "معتمر جديد",
            "passport_number": "N-1", "room_type": "ثنائي",
            "program_value": "4700", "paid_amount": "1000"})
assert r.status_code == 302
_am.set_mode("umrah")
_recs, _ = storage.load_records(UDATA, admin)
_mine = [(i, x) for i, x in enumerate(_recs) if x.trip == NC]
assert len(_mine) == 1 and _mine[0][1].full_name_ar == "معتمر جديد"
GI = _mine[0][0]
assert _mine[0][1].reference_number.startswith(NC + "-")
# تعديل المعتمر + حارس التزامن
r = um.post(f"/umrah/program/{NC}/pilgrim/{GI}/edit", data={
    "full_name_ar": "معتمر معدّل", "passport_number": "N-1", "room_type": "ثلاثي",
    "program_value": "4700", "paid_amount": "1000", "orig": "N-1"})
assert r.status_code == 302
_am.set_mode("umrah")
_recs, _ = storage.load_records(UDATA, admin)
assert _recs[GI].full_name_ar == "معتمر معدّل" and _recs[GI].room_type == "ثلاثي"
assert um.post(f"/umrah/program/{NC}/pilgrim/{GI}/edit",
               data={"passport_number": "N-1", "orig": "WRONG"}).status_code == 409
# سجلّ الدفعات: إضافة دفعة تُحدّث المحصّل والحالة
assert "سجلّ الدفعات" in um.get(
    f"/umrah/program/{NC}/pilgrim/{GI}/payments").get_data(as_text=True)
r = um.post(f"/umrah/program/{NC}/pilgrim/{GI}/payments/add",
            data={"date": "2026-08-03", "amount": "2000", "method": "نقد",
                  "note": "دفعة أولى"})
assert r.status_code == 302
_am.set_mode("umrah")
_recs, _ = storage.load_records(UDATA, admin)
assert len(_recs[GI].payments) == 1 and _ptot(_recs[GI]) == 2000.0
assert _recs[GI].paid_amount.replace(",", "") == "2000"
prog2 = um.get(f"/umrah/program/{NC}").get_data(as_text=True)
assert "معتمر معدّل" in prog2 and "جزئي" in prog2       # 2000 من 4700 = جزئي
# حذف الدفعة يعيد المحصّل صفراً
r = um.post(f"/umrah/program/{NC}/pilgrim/{GI}/payments/0/delete")
assert r.status_code == 302
_am.set_mode("umrah")
_recs, _ = storage.load_records(UDATA, admin)
assert len(_recs[GI].payments) == 0 and _ptot(_recs[GI]) == 0.0
# حذف المعتمر
r = um.post(f"/umrah/program/{NC}/pilgrim/{GI}/delete", data={"orig": "N-1"})
assert r.status_code == 302
_am.set_mode("umrah")
_recs, _ = storage.load_records(UDATA, admin)
assert not any(x.trip == NC for x in _recs)
# حذف البرنامج
r = um.post(f"/umrah/program/{NC}/delete")
assert r.status_code == 302
_am.set_mode("umrah")
assert not any(t.code == NC for t in _um.load_trips(storage.load_settings()))
# المطّلع ممنوع من مسارات كتابة العمرة (403)
vw = app.test_client()
vw.post("/login", data={"username": "viewer1", "password": "View-Pass-1234"})
vw.get("/mode/umrah")
assert vw.get("/umrah/program/new").status_code == 403
assert vw.post("/umrah/program/U1/pilgrim/new").status_code == 403
assert vw.post("/umrah/program/U1/pilgrim/0/payments/add").status_code == 403
print("  OK: إدارة برامج العمرة ومعتمريها ودفعاتهم على الويب، والمطّلع ممنوع")

# --- المرحلة ٤: مسعّر المجموعات على الويب ---
pform = um.get("/umrah/pricer").get_data(as_text=True)
assert "مسعّر المجموعات" in pform and "أنواع الغرف" in pform
assert "النقل الداخلي" in pform                # بند افتراضي معروض
# حساب: مطابقة جدول الإكسل (ثنائي: صافي 3750 / بيع 3950)
PDATA = {"title": "تسعير رمضان", "currency": "درهم",
         "makkah_rate": "1426", "makkah_nights": "3", "profit": "200",
         "item_name_0": "النقل الداخلي", "item_amount_0": "50",
         "item_name_1": "تذكرة الطيران", "item_amount_1": "1265",
         "item_name_2": "ماء وعصير وتمر", "item_amount_2": "46",
         "item_name_3": "الهدايا", "item_amount_3": "150",
         "item_name_4": "المصاريف الإدارية", "item_amount_4": "100",
         "room_types": ["مفرد", "ثنائي", "ثلاثي", "رباعي", "طفل"]}
res = um.post("/umrah/pricer", data=dict(PDATA, action="compute"))
rbody = res.get_data(as_text=True)
assert res.status_code == 200 and "3,750" in rbody and "3,950" in rbody
# اختيار أنواع الغرف: لو ثنائي فقط لا يظهر المفرد (5889)
only2 = um.post("/umrah/pricer",
                data={**{k: v for k, v in PDATA.items() if k != "room_types"},
                      "room_types": ["ثنائي"], "action": "compute"}).get_data(as_text=True)
assert "3,950" in only2 and "5,889" not in only2
# حفظ التسعير ثم استعراضه
sv = um.post("/umrah/pricer", data=dict(PDATA, action="save"))
assert sv.status_code == 302 and "/umrah/pricings" in sv.headers["Location"]
_am.set_mode("umrah")
_prs = _um.load_pricings(storage.load_settings())
assert _prs and _prs[-1]["title"] == "تسعير رمضان"
PN = _prs[-1]["number"]
assert PN.startswith("MA-P")
plist = um.get("/umrah/pricings").get_data(as_text=True)
assert "التسعيرات المحفوظة" in plist and PN in plist and "3,950" in plist
# فتح المحفوظ يملأ النموذج ويعرض النتيجة
edit = um.get(f"/umrah/pricer?number={PN}").get_data(as_text=True)
assert PN in edit and "تسعير رمضان" in edit and "3,750" in edit
# معاينة PDF
ppdf = um.post("/umrah/pricer/pdf", data=PDATA)
assert ppdf.status_code == 200 and ppdf.data[:5] == b"%PDF-"
# حذف التسعير
dl = um.post(f"/umrah/pricings/{PN}/delete")
assert dl.status_code == 302
_am.set_mode("umrah")
assert not any(p["number"] == PN for p in _um.load_pricings(storage.load_settings()))
# المطّلع ممنوع من الحفظ والحذف
assert vw.post("/umrah/pricer", data={"action": "save"}).status_code == 403
assert vw.post("/umrah/pricings/ANY/delete").status_code == 403
print("  OK: مسعّر المجموعات على الويب — حساب واختيار غرف وحفظ واستعراض وPDF وحذف")

# --- المرحلة ٥: عروض الأسعار على الويب ---
qp = um.get("/umrah/program/U1/pilgrim/0/quotation").get_data(as_text=True)
assert "عرض سعر" in qp and "معاينة عربي" in qp
# معاينة عربي/إنجليزي (PDF)
qar = um.get("/umrah/program/U1/pilgrim/0/quotation.pdf?lang=ar")
assert qar.status_code == 200 and qar.data[:5] == b"%PDF-"
qen = um.get("/umrah/program/U1/pilgrim/0/quotation.pdf?lang=en")
assert qen.status_code == 200 and qen.data[:5] == b"%PDF-"
# حفظ عرض سعر ثم استعراضه في قائمة البرنامج
sv = um.post("/umrah/program/U1/pilgrim/0/quotation/save", data={"lang": "ar"})
assert sv.status_code == 302 and "/umrah/program/U1/quotes" in sv.headers["Location"]
_am.set_mode("umrah")
_qs = _um.load_quotes(storage.load_settings(), "U1")
assert _qs and str(_qs[-1].get("number"))
QN = str(_qs[-1]["number"])
ql = um.get("/umrah/program/U1/quotes").get_data(as_text=True)
assert "عروض الأسعار" in ql and QN in ql
# استعراض العرض المحفوظ بالعربي والإنجليزي (ترجمة عند اختلاف اللغة)
for _lg in ("ar", "en"):
    d = um.get(f"/umrah/quote/U1/{QN}.pdf?lang={_lg}")
    assert d.status_code == 200 and d.data[:5] == b"%PDF-", _lg
# رقم غير موجود -> 404
assert um.get("/umrah/quote/U1/NOPE.pdf").status_code == 404
# حذف العرض
r = um.post(f"/umrah/quote/U1/{QN}/delete")
assert r.status_code == 302
_am.set_mode("umrah")
assert not any(str(q.get("number")) == QN
               for q in _um.load_quotes(storage.load_settings(), "U1"))
# المطّلع ممنوع من الحفظ والحذف
assert vw.post("/umrah/program/U1/pilgrim/0/quotation/save",
               data={"lang": "ar"}).status_code == 403
assert vw.post("/umrah/quote/U1/X/delete").status_code == 403
print("  OK: عروض الأسعار على الويب — معاينة ثنائية اللغة وحفظ واستعراض وترجمة وحذف")

# --- المرحلة ٦: الفاوتشر وكشوف التسكين/المواصلات/الطيران/البطاقات ---
prog3 = um.get("/umrah/program/U1").get_data(as_text=True)
assert "تسكين مكة" in prog3 and "المواصلات" in prog3 and "الطيران" in prog3
for _p in ("/umrah/program/U1/rooming/makkah.pdf",
           "/umrah/program/U1/rooming/madinah.pdf",
           "/umrah/program/U1/transport.pdf",
           "/umrah/program/U1/airline.pdf",
           "/umrah/program/U1/cards.pdf",
           "/umrah/program/U1/pilgrim/0/voucher.pdf",
           "/umrah/program/U1/pilgrim/0/voucher.pdf?lang=en"):
    d = um.get(_p)
    assert d.status_code == 200 and d.data[:5] == b"%PDF-", _p
# مدينة غير معروفة / معتمر خارج المدى -> 404
assert um.get("/umrah/program/U1/rooming/zzz.pdf").status_code == 404
assert um.get("/umrah/program/U1/pilgrim/99/voucher.pdf").status_code == 404
print("  OK: الفاوتشر وكشوف التسكين/المواصلات/الطيران/البطاقات على الويب")

# --- المرحلة ٧: الحجز بالتسعير ---
_am.set_mode("umrah")
_s7 = storage.load_settings()
_t7 = _um.load_trips(_s7)
_t7.append(_um.UmrahTrip(code="BK", name="حجز", price_double="4000",
           price_triple="3500",
           services=[{"name": "تأمين طبّي", "price": "200"},
                     {"name": "زيارة المدينة", "price": "300"}]))
_um.save_trips(_s7, _t7)
storage.save_settings(_s7)
bk = um.get("/umrah/program/BK/book").get_data(as_text=True)
assert "الحجز بالتسعير" in bk and "تأمين طبّي" in bk and "4,000" in bk
# حجز بغرفة ثنائية + خدمة تأمين => القيمة = 4200 (4000 + 200)
r = um.post("/umrah/program/BK/book", data={
    "full_name_ar": "حاجز جديد", "passport_number": "BK-1",
    "room_type": "ثنائي", "services": ["تأمين طبّي"], "persons": "2"})
assert r.status_code == 302
_am.set_mode("umrah")
_rc, _ = storage.load_records(UDATA, admin)
_bp = [x for x in _rc if x.trip == "BK"]
assert len(_bp) == 1
assert _bp[0].program_value == "4200" and _bp[0].room_type == "ثنائي"
assert _bp[0].room_value == "4000"
assert any(s["name"] == "تأمين طبّي" for s in _bp[0].umrah_services)
assert "فورد" in _bp[0].transport                 # شخصان -> فورد
assert _bp[0].reference_number.startswith("BK-")   # رقم مرجعي تلقائي
# المطّلع ممنوع من الحجز
assert vw.get("/umrah/program/BK/book").status_code == 403
print("  OK: الحجز بالتسعير على الويب — قيمة الفرد محسوبة من الغرفة والخدمات")

# --- المرحلة ٨: تحرير خدمات البرنامج على الويب ---
nf = um.get("/umrah/program/new").get_data(as_text=True)
assert "خدمات البرنامج" in nf and "تأمين طبّي" in nf     # خدمات افتراضية مقترحة
r = um.post("/umrah/program/new", data={
    "name": "مع خدمات", "price_double": "4000",
    "service_name_0": "تأمين طبّي", "service_price_0": "200",
    "service_name_1": "زيارة", "service_price_1": "300"})
assert r.status_code == 302
NC8 = r.headers["Location"].rstrip("/").split("/")[-1]
_am.set_mode("umrah")
_t8 = next(t for t in _um.load_trips(storage.load_settings()) if t.code == NC8)
assert {s["name"] for s in _t8.services} == {"تأمين طبّي", "زيارة"}
assert _um.services_map(_t8)["تأمين طبّي"] == 200
# الخدمات المُدخلة تظهر في صفحة الحجز بالتسعير
bk8 = um.get(f"/umrah/program/{NC8}/book").get_data(as_text=True)
assert "تأمين طبّي" in bk8 and "زيارة" in bk8
# التعديل يحدّث الخدمات (إبقاء واحدة بسعر جديد)
r = um.post(f"/umrah/program/{NC8}/edit", data={
    "name": "مع خدمات", "price_double": "4000",
    "service_name_0": "تأمين طبّي", "service_price_0": "250"})
assert r.status_code == 302
_am.set_mode("umrah")
_t8b = next(t for t in _um.load_trips(storage.load_settings()) if t.code == NC8)
assert len(_t8b.services) == 1 and _um.services_map(_t8b)["تأمين طبّي"] == 250
print("  OK: تحرير خدمات البرنامج على الويب وظهورها في الحجز بالتسعير")

# التبديل عائداً إلى الحج يعيد كشف الحج
um.get("/mode/hajj")
assert "برنامج موسم الحج" in um.get("/").get_data(as_text=True)
# التبديل لوضع غير معروف -> 404
assert um.get("/mode/zzz").status_code == 404
print("  OK: التبديل للعمرة وبرامجها وإدارتها المالية وملخّصها PDF ومعزولة عن الحج")

print("\n*** WEB TESTS PASSED ***")

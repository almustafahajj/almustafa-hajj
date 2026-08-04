# -*- coding: utf-8 -*-
"""اختبار برنامج العمرة: نموذج البرنامج، التسعير، النقل، الحجز، ونافذة المعتمرين."""
import sys, io
import os as _os
_HERE = _os.path.dirname(_os.path.abspath(__file__))
_ROOT = _os.path.dirname(_HERE)
_OUTDIR = _os.path.join(_HERE, "_out")
_os.makedirs(_OUTDIR, exist_ok=True)
sys.path.insert(0, _ROOT)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import pathlib
import tkinter as tk
from tkinter import messagebox as _mb
# تعطيل الرسائل المنبثقة الحاجبة أثناء الاختبار (تحاكي ضغط المستخدم)
_mb.showinfo = lambda *a, **k: "ok"
_mb.showwarning = lambda *a, **k: "ok"
_mb.showerror = lambda *a, **k: "ok"
_mb.askyesno = lambda *a, **k: True
import hajj_app.storage as st
from hajj_app import app_mode, umrah
from hajj_app.mrz import PassportData

WORK = pathlib.Path(_OUTDIR) / "umrah"
WORK.mkdir(parents=True, exist_ok=True)
for _f in WORK.glob("*"):
    if _f.is_file():
        _f.unlink()
app_mode.set_mode("umrah")
st.default_data_path = lambda: WORK / app_mode.data_filename()

print("=== النموذج والتخزين والترحيل ===")
s = {}
assert umrah.load_trips(s) == []
t1 = umrah.UmrahTrip(code="U1", name="رمضان", makkah_hotel="فندق",
                     price_double="4500",
                     services=[{"name": "تأمين طبّي", "price": "200"}])
umrah.save_trips(s, [t1, umrah.UmrahTrip(code="U2", name="شعبان")])
back = umrah.load_trips(s)
assert [t.code for t in back] == ["U1", "U2"]
assert back[0].services == [{"name": "تأمين طبّي", "price": "200"}]
assert umrah.next_code(back) == "U3"
# ترحيل الصيغة القديمة: services=list[str] و price المفرد القديم
old = umrah.trip_from_dict({"code": "X", "price": "3000",
                            "services": ["زيارة المدينة المنوّرة"]})
assert old.price_double == "3000"                       # رُحّل السعر القديم
assert old.services == [{"name": "زيارة المدينة المنوّرة", "price": ""}]
print("  OK: حفظ/تحميل + ترحيل الأسعار والخدمات القديمة")

print("\n=== التسعير والنقل ===")
t = umrah.UmrahTrip(code="P1", price_single="6000", price_double="4500",
                    price_triple="4000", price_quad="3500",
                    services=[{"name": "تأمين طبّي", "price": "200"},
                              {"name": "زيارة المدينة المنوّرة", "price": "300"}])
assert umrah.room_price(t, "price_double") == 4500
assert umrah.services_map(t)["تأمين طبّي"] == 200
assert umrah.package_per_person(t, "price_double", ["تأمين طبّي"]) == 4700
assert umrah.package_per_person(
    t, "price_single", ["تأمين طبّي", "زيارة المدينة المنوّرة"]) == 6500
assert "فورد" in umrah.suggest_transport(2)
assert "جيمس" in umrah.suggest_transport(3) and "جيمس" in umrah.suggest_transport(6)
assert umrah.suggest_transport(0) == ""
print("  OK: أسعار الغرف + الخدمات لكل شخص + قاعدة النقل حسب العدد")

print("\n=== ربط المعتمر بالبرنامج ===")
recs = [PassportData(trip="U1"), PassportData(trip="U2"), PassportData(trip="U1")]
assert len(umrah.trip_pilgrims(recs, "U1")) == 2
assert umrah.trip_pilgrims(recs, "U9") == []
print("  OK: ترشيح معتمري كل برنامج")

print("\n=== الواجهة: البرامج، الحجز بالتسعير، والمالية ===")
import hajj_app.umrah_gui as ug
root = tk.Tk(); root.withdraw()
app = ug.UmrahApp(root, session=None)
tt = umrah.UmrahTrip(code="P1", name="رمضان ١", price_triple="4000",
                     services=[{"name": "تأمين طبّي", "price": "200"}])
app._on_trip_saved(tt)
assert app.tree.get_children() == ("P1",)
tt.capacity = "3"                                  # سعة محدّدة للاختبار
tt.airline = "الطيران السعودي"; tt.flight_out = "SV553"
tt.depart_date = "2026-03-01"; tt.makkah_hotel = "كونراد"
win = ug.TripPilgrimsWindow(app, tt)
# حجز بالتسعير: 3 أشخاص، غرفة ثلاثية، + خدمة تأمين (مسعّرة)
bd = ug.BookingDialog(win, tt)
bd.persons.set("3"); bd.room.set("ثلاثي")
for _r in bd.svc_rows:                              # اختيار خدمة تأمين طبّي
    if _r["name"] == "تأمين طبّي":
        _r["on"].set(True)
bd._recalc()
assert getattr(bd, "_per_person") == 4200          # 4000 + 200
assert "جيمس" in bd.transport.get()                # النقل تلقائي (3 أشخاص)
# إضافة خدمة مخصّصة مسعّرة يدوياً تزيد سعر الفرد
bd._new_svc.set("خدمة خاصة"); bd._new_price.set("500")
bd._add_custom_service()
assert bd._per_person == 4700                       # 4200 + 500
bd.svc_rows[-1]["on"].set(False); bd._recalc()      # نلغيها لبقية الاختبار
assert bd._per_person == 4200
# التسعير + السفر/الإقامة من البرنامج + الرقم المرجعي التلقائي لكل معتمر
for nm in ("محمد", "أحمد", "سالم"):
    r = PassportData(full_name_ar=nm)
    bd._apply_booking(r)
    assert r.program_value == "4200" and r.room_type == "ثلاثي"    # 4000 + 200
    assert r.room_value == "4000"                                   # الأساس
    assert any(s["name"] == "تأمين طبّي" for s in r.umrah_services)  # خدمة مسعّرة
    assert "جيمس" in r.transport and r.trip == "P1"
    assert r.airline == "الطيران السعودي" and r.hotel == "كونراد"   # من البرنامج
    assert r.reference_number.startswith("P1-")                     # تلقائي
    bd._commit_person(r)
assert bd._added == 3
refs = [r.reference_number for r in umrah.trip_pilgrims(app.records, "P1")]
assert len(set(refs)) == 3                          # أرقام مرجعية فريدة
recs2, _ = st.load_records()
assert len(recs2) == 3 and all(r.program_value == "4200" for r in recs2)
fin = win.fin.cget("text")
assert "العدد: 3" in fin and "المقاعد المتبقّية: 0 من 3" in fin
# السعة: لا تقبل أكثر من المحدَّد
assert win._seats_left() == 0 and win._check_capacity(1) is False
# رمز البرنامج يُذكر في الكشف
assert win._prog_label() == "P1 — رمضان ١"
# الموسم = سنة ميلادية: تصفية البرامج حسب سنة المغادرة
app._season.set("2026")
app.trips += [umrah.UmrahTrip(code="A", name="أ", depart_date="2026-03-01"),
              umrah.UmrahTrip(code="B", name="ب", depart_date="2027-04-01")]
app._reload()
c26 = set(app.tree.get_children())
assert "A" in c26 and "B" not in c26 and "P1" in c26      # P1 مغادرته 2026
app._season.set("2027"); app._reload()
c27 = set(app.tree.get_children())
assert "B" in c27 and "A" not in c27 and "P1" not in c27
assert "١ يناير" in app._season_text() and "2027" in app._season_text()
# واجهة مبسّطة: الأدوات مجمّعة في قوائم منسدلة (Menubutton + Menu)
_ulabels = set()
for _m in getattr(app, "_menus", []):
    try:
        _end = _m.index("end")
    except Exception:
        _end = None
    for _i in range(0 if _end is None else _end + 1):
        try:
            _ulabels.add(_m.entrycget(_i, "label"))
        except Exception:
            pass
_utxt = " ".join(_ulabels)
assert len(app._menus) == 3                                  # ثلاث قوائم مبسّطة
for _need in ("تعديل البرنامج", "حذف البرنامج", "عروض الأسعار المحفوظة",
              "عرض سعر يدوي", "العروض اليدوية", "مسعّر المجموعات",
              "التسعيرات المحفوظة", "فاوتشر فندق يدوي"):
    assert _need in _utxt, _need
root.destroy()
print("  OK: التسعير والسفر والرقم المرجعي، السعة، رمز البرنامج، وموسم السنة")

print("\n=== نافذة التعديل في وضع العمرة ===")
import hajj_app.gui as _g
r2 = tk.Tk(); r2.withdraw()
dlg = _g.EditDialog(r2, PassportData(), lambda _r: None, umrah=True)
# مُستبعَد: حقول الحج + السفر/الفندق (من البرنامج) + الصحة والطوارئ
for gone in ("program", "group", "visa_number", "permit_status", "hady",
             "masar_number", "executive_service", "airline", "flight_number",
             "hotel", "arrival_date", "blood_type", "emergency_name"):
    assert gone not in dlg.vars, gone
# موجود: البيانات + الرقم المرجعي + المالية مع تاريخ الدفع وطريقته
for keep in ("full_name_ar", "passport_number", "reference_number",
             "program_value", "paid_amount", "payment_date", "payment_method",
             "room_type", "transport"):
    assert keep in dlg.vars, keep
dlg.destroy()
# وضع الحج يُبقي كل الحقول
dlg2 = _g.EditDialog(r2, PassportData(), lambda _r: None)
assert "visa_number" in dlg2.vars and "program" in dlg2.vars and "hotel" in dlg2.vars
dlg2.destroy()
r2.destroy()
print("  OK: العمرة تُلغي الحج/السفر/الصحة، وتضيف تاريخ/طريقة الدفع، والحج يُبقيها")

print("\n=== الخدمات الجديدة وكشف المعتمرين ومسمّياته ===")
app_mode.set_mode("umrah")
for extra in ("المطوّف", "خدمة التنفيذي في الاستقبال",
              "خدمة التنفيذي في المغادرة", "خدمة مُرافق"):
    assert extra in umrah.DEFAULT_SERVICES, extra
assert app_mode.label("release_name") == "ميسّر العمرة"      # اسم التطبيق
# صفّ الكشف بالأعمدة المطلوبة (الموظف المسؤول من خانة staff، آخر عمود)
rec = PassportData(full_name_ar="سعيد", passport_number="A9",
                   expiry_date="2030-01-01", nationality_ar="الإمارات",
                   hotel="كونراد", room_type="ثنائي", airline="السعودية",
                   program_value="5000", paid_amount="2000", staff="أيمن الشهابي")
row = umrah.report_row(rec, 1, "رمضان")
assert [k for k, _l in umrah.REPORT_COLUMNS] == [
    "serial", "full_name_ar", "family_number", "passport_number", "expiry_date",
    "nationality_ar", "program", "hotel", "room_type", "airline",
    "program_value", "paid_amount", "remaining"]              # بلا staff (لا إكسل/غرف)
# الموظف المسؤول عمود إضافي للمعاينة والملخّص المالي فقط
assert umrah.REPORT_STAFF_COLUMN == ("staff", "الموظف المسؤول")
assert row["staff"] == "أيمن الشهابي"                          # من بيانات المعتمر
# الجنسية قبل البرنامج، ورقم العائلة مذكور
_keys = [k for k, _l in umrah.REPORT_COLUMNS]
assert _keys.index("nationality_ar") < _keys.index("program")
assert "family_number" in _keys
assert row["program"] == "رمضان" and row["remaining"] == "3,000"
# تصدير PDF + إكسل بمسمّيات العمرة (لا «حاج» في الكشف)
from hajj_app.pdf_io import export_umrah_pdf
from hajj_app.excel_io import export_umrah_excel
pdf = WORK / "kashf.pdf"
export_umrah_pdf([rec], pdf, program_name="رمضان")
assert pdf.read_bytes()[:5] == b"%PDF-" and pdf.stat().st_size > 3000
xlsx = WORK / "kashf.xlsx"
export_umrah_excel([rec], xlsx, program_name="رمضان")
from openpyxl import load_workbook
ws = load_workbook(xlsx).active
_ncols = len(umrah.REPORT_COLUMNS)
assert [ws.cell(row=2, column=c).value for c in range(1, _ncols + 1)] == \
    [lbl for _k, lbl in umrah.REPORT_COLUMNS]
cells = [str(c.value) for r in ws.iter_rows() for c in r if c.value is not None]
assert not any("حاج" in t for t in cells), "بقيت كلمة حاج في الكشف"
assert any("المعتمرين" in t for t in cells)                # عنوان الكشف
# الموظف المسؤول لا يظهر في إكسل (المعاينة والملخّص المالي فقط)
assert not any("أيمن الشهابي" in t for t in cells)
assert "الموظف المسؤول" not in [lbl for _k, lbl in umrah.REPORT_COLUMNS]
print("  OK: خدمات جديدة + كشف بأعمدته ومسمّيات العمرة (بلا «حاج») + اسم التطبيق")

print("\n=== التسكين (مكة/المدينة) ===")
app_mode.set_mode("umrah")
rr = [PassportData(full_name_ar=f"م{i}", passport_number=f"P{i}", room_type=rt,
                   trip="R1") for i, rt in enumerate(
                       ["ثنائي", "ثنائي", "ثنائي", "مفرد"])]
assert umrah.auto_assign_rooms(rr, "makkah_room") == (3, 0)  # 3 ثنائي→غرفتان + مفرد
rooms, un = umrah.rooming_rooms(rr, "makkah_room")
assert len(rooms) == 3 and not un
assert all(not r.madinah_room for r in rr)                  # المدينة مستقلة عن مكة
umrah.auto_assign_rooms(rr, "madinah_room")
assert all(r.madinah_room and r.makkah_room for r in rr)    # الحقلان مستقلان
from hajj_app.pdf_io import export_umrah_rooming_pdf
p = WORK / "rooms.pdf"
export_umrah_rooming_pdf(rr, p, city_label="مكة المكرّمة", hotel="كونراد",
                         nights="7", program_name="R1 — رمضان",
                         room_field="makkah_room")
assert p.read_bytes()[:5] == b"%PDF-" and p.stat().st_size > 2500
# نافذة التسكين: توزيع، عرض، ومسح
r3 = tk.Tk(); r3.withdraw()
app3 = ug.UmrahApp(r3, session=None)
t3 = umrah.UmrahTrip(code="R1", name="رمضان", makkah_hotel="كونراد",
                     madinah_hotel="دار التقوى")
app3.trips.append(t3)
app3.records.extend([PassportData(full_name_ar=f"س{i}", room_type="ثنائي",
                                  trip="R1") for i in range(4)])
rw = ug.RoomingWindow(app3, t3)
rw._auto("makkah")
assert all(r.makkah_room for r in umrah.trip_pilgrims(app3.records, "R1"))
vals = [rw._trees["makkah"].item(i, "values")
        for i in rw._trees["makkah"].get_children()]
assert all(v[4] not in ("", "—") for v in vals)            # رقم الغرفة ظاهر
rw._clear("makkah")                                        # يؤكّد (مُثبَّت) ويمسح
assert all(not r.makkah_room for r in umrah.trip_pilgrims(app3.records, "R1"))
r3.destroy()
print("  OK: توزيع تلقائي مستقل لكل مدينة + كشف الغرف + نافذة التسكين")

print("\n=== سعر الطفل + خدمات المعتمر + المواصلات والطيران ===")
app_mode.set_mode("umrah")
# سعر الطفل (بدون سرير) ضمن أنواع الغرف، بلا سعة
assert ("price_child", "طفل (بدون سرير)", 0) in umrah.ROOM_TYPES
assert umrah.room_capacity_of("طفل (بدون سرير)") == 0
kids = [PassportData(room_type="طفل (بدون سرير)"), PassportData(room_type="ثنائي"),
        PassportData(room_type="ثنائي")]
umrah.auto_assign_rooms(kids, "makkah_room")
assert kids[0].makkah_room == "" and kids[1].makkah_room     # الطفل بلا غرفة
# المركبات: فورد ≤ شخصين، جيمس حتى ٦ — 8 أشخاص = جيمس(٦) + فورد(٢)
vv = [PassportData() for _ in range(8)]
assert umrah.auto_assign_vehicles(vv) == 2
assert vv[0].vehicle == "جيمس 1" and vv[7].vehicle == "فورد 2"
two = [PassportData(), PassportData()]
assert umrah.auto_assign_vehicles(two) == 1 and two[0].vehicle.startswith("فورد")
# خدمات المعتمر في نافذة التعديل: القيمة = الأساس + الخدمات
import hajj_app.gui as _g
r4 = tk.Tk(); r4.withdraw()
trip4 = umrah.UmrahTrip(code="U1", services=[{"name": "تأمين طبّي", "price": "200"},
                                             {"name": "المطوّف", "price": "150"}])
rec4 = PassportData(full_name_ar="سعيد", room_value="4000", program_value="4000")
d4 = _g.EditDialog(r4, rec4, lambda _r: None, umrah=True, trip=trip4)
d4._svc_widgets["تأمين طبّي"][0].set(True)
d4._recalc_services()
assert d4.vars["program_value"].get() == "4,200"            # 4000 + 200 حيّاً
d4._save()
assert rec4.program_value == "4,200" and rec4.room_value == "4,000"
assert len(rec4.umrah_services) == 1
r4.destroy()
# كشوف PDF: المواصلات والطيران
from hajj_app.pdf_io import export_umrah_transport_pdf, export_airline_pdf
tr = [PassportData(full_name_ar=f"م{i}", passport_number=f"P{i}", vehicle="جيمس 1")
      for i in range(3)]
pt = WORK / "trans.pdf"
export_umrah_transport_pdf(tr, pt, program_name="U1 — رمضان")
assert pt.read_bytes()[:5] == b"%PDF-" and pt.stat().st_size > 2500
pf = WORK / "flight.pdf"
export_airline_pdf(tr, pf, title="Flight Manifest — U1")
assert pf.read_bytes()[:5] == b"%PDF-"
# نافذة المواصلات: توزيع تلقائي
import hajj_app.umrah_gui as _ug
r5 = tk.Tk(); r5.withdraw()
app5 = _ug.UmrahApp(r5, session=None)
t5 = umrah.UmrahTrip(code="U1", name="رمضان")
app5.trips.append(t5)
app5.records.extend([PassportData(full_name_ar=f"س{i}", trip="U1") for i in range(4)])
vw = _ug.TransportWindow(app5, t5)
vw._auto()
assert all(r.vehicle for r in umrah.trip_pilgrims(app5.records, "U1"))
r5.destroy()
print("  OK: سعر الطفل + خدمات المعتمر المسعّرة + المواصلات (فورد/جيمس) + الطيران")

print("\n=== الرضيع + PNR + سعة الغرف + كشف المواصلات بالفندق ===")
# الرضيع ضمن أنواع الغرف بلا سعة
assert ("price_infant", "رضيع", 0) in umrah.ROOM_TYPES
assert umrah.room_capacity_of("رضيع") == 0
# PNR الطيران يُورَّث للمعتمر من البرنامج
tp = umrah.UmrahTrip(code="U1", flight_pnr="ABC123", makkah_hotel="كونراد")
rp = PassportData()
umrah.apply_trip_to_record(tp, rp)
assert rp.pnr == "ABC123"
# سعة الغرف: التوزيع يتجاوز عدد الغرف المتاحة → تنبيه (n > available)
r6 = tk.Tk(); r6.withdraw()
app6 = _ug.UmrahApp(r6, session=None)
t6 = umrah.UmrahTrip(code="RM", name="رمضان", makkah_hotel="كونراد",
                     makkah_rooms="1", madinah_hotel="دار التقوى")
app6.trips.append(t6)
app6.records.extend([PassportData(full_name_ar=f"م{i}", room_type="مفرد",
                                  trip="RM") for i in range(3)])
rw6 = _ug.RoomingWindow(app6, t6)
assert rw6._available("makkah") == 1
# مع تحديد السعة: لا يُتجاوز عدد الغرف؛ يبقى فائض بلا غرفة
n_rooms, overflow = umrah.auto_assign_rooms(
    umrah.trip_pilgrims(app6.records, "RM"), "makkah_room", max_rooms=1)
assert n_rooms == 1 and overflow == 2       # غرفة واحدة فقط، واثنان بلا غرفة
r6.destroy()
# كشف المواصلات: يعرض الفندق ويوضّح الاشتراك (أكثر من راكب)
from hajj_app.pdf_io import export_umrah_transport_pdf
shared = [PassportData(full_name_ar=f"س{i}", passport_number=f"P{i}",
                       hotel="كونراد", vehicle="جيمس 1") for i in range(3)]
pv6 = WORK / "trans2.pdf"
export_umrah_transport_pdf(shared, pv6, program_name="U1", transport_pnr="TR9")
assert pv6.read_bytes()[:5] == b"%PDF-" and pv6.stat().st_size > 2500
print("  OK: الرضيع + توريث PNR + تنبيه سعة الغرف + كشف مواصلات بالفندق")

print("\n=== رابط الدفع + بذر رقم الغرفة + المالية والبطاقات ===")
app_mode.set_mode("umrah")
# «رابط دفع» ضمن طرق الدفع
import hajj_app.gui as _g2
assert "رابط دفع" in _g2.EditDialog.CHOICE_FIELDS["payment_method"]
# التسكين يأخذ رقم الغرفة من «الإقامة والحجز» إن وُجد
r7 = tk.Tk(); r7.withdraw()
app7 = _ug.UmrahApp(r7, session=None)
t7 = umrah.UmrahTrip(code="SD", name="رمضان", makkah_hotel="كونراد")
app7.trips.append(t7)
app7.records.append(PassportData(full_name_ar="سعيد", room_number="512", trip="SD"))
_ug.RoomingWindow(app7, t7)
rec7 = umrah.trip_pilgrims(app7.records, "SD")[0]
assert rec7.makkah_room == "512" and rec7.madinah_room == "512"   # مأخوذ تلقائياً
r7.destroy()
# الملخّص المالي + بطاقات العمرة PDF
from hajj_app.pdf_io import export_umrah_finance_pdf, export_umrah_cards_pdf
fr = [PassportData(full_name_ar=f"م{i}", passport_number=f"A{i}",
                   nationality_ar="الإمارات", hotel="كونراد", room_type="ثنائي",
                   airline="السعودية", pnr="P1", reference_number=f"U1-00{i}",
                   program_value="5000", paid_amount=str(2000 * (i % 3)),
                   payment_method="رابط دفع") for i in range(1, 5)]
pf = WORK / "fin.pdf"
export_umrah_finance_pdf(fr, pf, program_name="U1 — رمضان")
assert pf.read_bytes()[:5] == b"%PDF-" and pf.stat().st_size > 3000
pc = WORK / "cards.pdf"
export_umrah_cards_pdf(fr, pc, program_name="U1 — رمضان",
                       emergency_uae="+971500000000", emergency_ksa="+966500000000")
assert pc.read_bytes()[:5] == b"%PDF-" and pc.stat().st_size > 3000
# تنبيه صلاحية الجواز أقل من ٦ أشهر من تاريخ السفر + المنتهي
assert umrah.passport_expiry_soon(
    PassportData(expiry_date="2026-05-01"), "2026-03-01") is True
assert umrah.passport_expiry_soon(
    PassportData(expiry_date="2027-01-01"), "2026-03-01") is False
assert umrah.passport_expired(PassportData(expiry_date="2020-01-01")) is True
assert umrah.passport_expired(PassportData(expiry_date="2030-01-01")) is False
assert umrah.passport_flag(PassportData(expiry_date="2020-01-01"), "") is True
# الكشف الرئيسي: المتبقّي = السعة − المعتمرين
r8 = tk.Tk(); r8.withdraw()
app8 = _ug.UmrahApp(r8, session=None)
app8.trips.append(umrah.UmrahTrip(code="SE", name="س", capacity="5"))
app8.records.extend([PassportData(trip="SE") for _ in range(2)])
app8._reload()
assert str(app8.tree.item("SE", "values")[-1]) == "3"      # 5 − 2 مقاعد متبقّية
r8.destroy()
print("  OK: رابط الدفع + الغرفة + المالية + البطاقات + تنبيه/انتهاء الجواز + المقاعد")

print("\n=== نافذة الإدارة المالية (الدفعات والحالة والملخّص) ===")
from hajj_app.fields import sync_paid_amount, payment_total
r9 = tk.Tk(); r9.withdraw()
app9 = _ug.UmrahApp(r9, session=None)
t9 = umrah.UmrahTrip(code="FN", name="مالية")
app9.trips.append(t9)
app9.records.extend([
    PassportData(full_name_ar="مسدّد", trip="FN", room_type="ثنائي",
                 program_value="5000", paid_amount="5000"),
    PassportData(full_name_ar="جزئي", trip="FN", room_type="ثلاثي",
                 program_value="4000", paid_amount="1500"),
    PassportData(full_name_ar="غير مدفوع", trip="FN", room_type="مفرد",
                 program_value="6000", paid_amount="0"),
])
fw = _ug.UmrahFinanceWindow(app9.root, app9, t9)
# بطاقات الملخّص: القيمة/المحصّل/المتبقّي/النسبة/المتأخّرون
assert fw._card_vars["value"].get().replace(",", "") == "15000"
assert fw._card_vars["paid"].get().replace(",", "") == "6500"
assert fw._card_vars["remaining"].get().replace(",", "") == "8500"
assert fw._card_vars["pct"].get() == "43%"
assert fw._card_vars["owe"].get() == "2"
# حالات الصفوف (المسدّد/الجزئي/غير المدفوع) بألوانها
_st_by_name = {fw.tree.item(i, "values")[1]: fw.tree.item(i, "values")[6]
               for i in fw.tree.get_children()}
assert _st_by_name["مسدّد"] == "مسدّد"
assert _st_by_name["جزئي"] == "جزئي"
assert _st_by_name["غير مدفوع"] == "غير مدفوع"
_tags = {fw.tree.item(i, "values")[1]: fw.tree.item(i, "tags")[0]
         for i in fw.tree.get_children()}
assert _tags["مسدّد"] == "paid" and _tags["جزئي"] == "partial"
assert _tags["غير مدفوع"] == "unpaid"
# تسجيل دفعة عبر سجلّ الأقساط يُحدّث المحصّل والحالة
_rec_un = umrah.trip_pilgrims(app9.records, "FN")[2]     # «غير مدفوع»
_rec_un.payments.append({"date": "2026-08-03", "amount": "6000",
                         "method": "تحويل بنكي", "note": "كامل"})
sync_paid_amount(_rec_un)
assert payment_total(_rec_un) == 6000.0
fw._reload()
assert fw._card_vars["paid"].get().replace(",", "") == "12500"
assert fw._card_vars["owe"].get() == "1"                 # بقي «الجزئي» فقط
r9.destroy()
print("  OK: الإدارة المالية — بطاقات وحالات ودفعات (أقساط) وتحصيل حيّ")

print("\n=== مستندات العمرة: سند قبض، فاتورة، وعقد ===")
app_mode.set_mode("umrah")
from hajj_app.pdf_io import (export_umrah_receipt_pdf, export_umrah_invoice_pdf,
                             export_umrah_contract_pdf)
recd = PassportData(full_name_ar="عبدالرحمن", passport_number="A1", hotel="كونراد",
                    room_type="مفرد", program_value="7400", paid_amount="5000",
                    umrah_services=[{"name": "تأمين طبّي", "price": "200"}])
for _fn, _nm in [(export_umrah_receipt_pdf, "r"), (export_umrah_invoice_pdf, "i"),
                 (export_umrah_contract_pdf, "c")]:
    _pp = WORK / f"doc_{_nm}.pdf"
    _fn(recd, _pp, program_name="ديسمبر")
    assert _pp.read_bytes()[:5] == b"%PDF-" and _pp.stat().st_size > 3000
# أزرار المستندات للمعتمر المحدّد في نافذة المعتمرين
import hajj_app.gui as _g3
_g3.open_preview = lambda parent, fn, name, ext: (fn(str(WORK / "sel.pdf")), "x")[1]
r9 = tk.Tk(); r9.withdraw()
app9 = _ug.UmrahApp(r9, session=None)
t9 = umrah.UmrahTrip(code="D1", name="ديسمبر")
app9.trips.append(t9)
app9.records.append(PassportData(full_name_ar="محمد", trip="D1",
                                 program_value="5000", paid_amount="2000"))
# فاوتشر الفندق (إقامات مكة/المدينة بتواريخ متسلسلة)
from hajj_app.pdf_io import export_umrah_voucher_pdf
tripv = umrah.UmrahTrip(code="U1", name="ديسمبر", makkah_hotel="جميرا مكة",
                        makkah_nights="4", madinah_hotel="دار التقوى",
                        madinah_nights="3", depart_date="2026-08-07",
                        transport="جيمس")
pv = WORK / "voucher.pdf"
_recv = PassportData(full_name_ar="خالد", full_name_en="Khalid",
                     passport_number="A1", room_type="ثنائي")
export_umrah_voucher_pdf(_recv, pv, trip=tripv, program_name="ديسمبر")
assert pv.read_bytes()[:5] == b"%PDF-" and pv.stat().st_size > 3000
# صفحة عرضية (Landscape) تتّسع في صفحة واحدة، بالعربية والإنجليزية
import fitz as _fitz
for _lg in ("ar", "en"):
    _pl = WORK / f"voucher_{_lg}.pdf"
    export_umrah_voucher_pdf(_recv, _pl, trip=tripv, program_name="ديسمبر",
                             lang=_lg)
    _pd = _fitz.open(str(_pl))
    assert _pd.page_count == 1, f"الفاوتشر ({_lg}) تجاوز صفحة واحدة"
    assert _pd[0].rect.width > _pd[0].rect.height, "الصفحة ليست عرضية"
    _pd.close()
w9 = _ug.TripPilgrimsWindow(app9, t9)
w9.tree.selection_set("0")
for _do in (w9.do_receipt, w9.do_invoice, w9.do_contract):
    _do()
    assert (WORK / "sel.pdf").read_bytes()[:5] == b"%PDF-"
# فاوتشر الفندق يفتح محرّراً قابلاً للتعديل (رقم تسلسلي، إطلالة/نقل منسدلة،
# تاريخ منسدل، إضافة/حذف خلايا) قبل المعاينة
w9.do_voucher()
_vdlg = [w for w in app9.root.winfo_children()
         if isinstance(w, _ug.VoucherEditorDialog)] or \
        [w for w in w9.winfo_children()
         if isinstance(w, _ug.VoucherEditorDialog)]
assert _vdlg, "محرّر الفاوتشر لم يُفتح"
_ed = _vdlg[-1]
assert _ed._number.startswith("MA")          # رقم تسلسلي تلقائي
_n0 = len(_ed._stay_rows)
_ed._add_stay_row(["المدينة المنوّرة", "دار الإيمان", "ثلاثي", "Haram",
                   "", "", "2", "إفطار"])
assert len(_ed._stay_rows) == _n0 + 1
_ed._del_row(_ed._stay_rows, _ed._stay_rows[-1])
assert len(_ed._stay_rows) == _n0
_ed._add_transport_row(["FORD", "2026", "استقبال من المطار"])
_ed._add_contact_row(["مشرف", "سالم", "+966 55 000 0000"])
_data = _ed._collect()
assert _data["contacts"] and _data["terms"] and _data["transport_rows"]
assert _data["number"] == _ed._number and "-" in _data["date"]
assert "title_ar" not in _data          # العنوان محذوف من التعديل
_ed._preview()
assert (WORK / "sel.pdf").read_bytes()[:5] == b"%PDF-"
_ed.destroy()
# طلب حجز مواصلات: خطاب لشركة النقل (بيانات الضيف + جدولا الطيران والحركة)
from hajj_app.pdf_io import (build_transport_request_data,
                             export_umrah_transport_request_pdf)
_trb = build_transport_request_data(_recv, trip=tripv, program_name="ديسمبر",
                                    number="MA-T0001", recipient="جنى")
assert _trb["guest_ar"] == "خالد" and _trb["recipient"] == "جنى"
assert len(_trb["movements"]) == 4 and len(_trb["flights"]) == 2
assert "جميرا مكة" in _trb["reservations"]
_ptr = WORK / "treq.pdf"
export_umrah_transport_request_pdf(_recv, _ptr, data=_trb)
assert _ptr.read_bytes()[:5] == b"%PDF-" and _ptr.stat().st_size > 3000
# المحرّر: يفتح من نافذة المعتمرين ويعاين بعد التعديل
w9.do_transport_request()
_tr = [w for w in w9.winfo_children()
       if isinstance(w, _ug.TransportRequestEditorDialog)] or \
      [w for w in app9.root.winfo_children()
       if isinstance(w, _ug.TransportRequestEditorDialog)]
assert _tr, "محرّر طلب المواصلات لم يُفتح"
_trd = _tr[-1]
assert _trd._number.startswith("MA") and _trd._meta["guest_ar"].get()
_trd._meta["recipient"].set("جنى عبد الله السكيت")
_nm = len(_trd._move_rows)
_trd._add_row(["2026-08-09", "من الفندق إلى المطار", "1", "FORD", "2026",
               "10:00"], _ug.TREQ_MOVE_HEADS, (11, 22, 5, 10, 8, 8),
              _trd._move_rows, _trd._move_box)
assert len(_trd._move_rows) == _nm + 1
_trc = _trd._collect()
assert _trc["recipient"] == "جنى عبد الله السكيت" and _trc["movements"]
assert _trc["office_manager"] == "أيمن الشهابي"
_trd._preview()
assert (WORK / "sel.pdf").read_bytes()[:5] == b"%PDF-"
_trd.destroy()
# فاوتشر يدوي لأي حجز خارج البرامج (rec/trip فارغان)
app9.new_manual_voucher()
_mv = [w for w in r9.winfo_children()
       if isinstance(w, _ug.VoucherEditorDialog)][-1]
assert _mv.trip is None and _mv._number.startswith("MA")
_mv._add_stay_row(["مكة المكرّمة", "فندق", "ثنائي", "Kaaba",
                   "2026-08-07", "2026-08-11", "4", "إفطار"])
_mv._preview()
assert (WORK / "sel.pdf").read_bytes()[:5] == b"%PDF-"
_mv.destroy()
r9.destroy()
print("  OK: سند وفاتورة وعقد وفاوتشر الفندق وطلب المواصلات لكل معتمر")

# === عرض السعر (Quotation) ===
from hajj_app.pdf_io import build_quotation_data, export_umrah_quotation_pdf
_tripq = umrah.UmrahTrip(code="Q1", name="سبتمبر", depart_date="2026-09-04",
                         return_date="2026-09-11", makkah_hotel="جميرا مكة",
                         makkah_nights="4", madinah_hotel="دار الإيمان",
                         madinah_nights="3", airline="FLYDUBAI",
                         out_depart_time="14:25", ret_depart_time="17:45",
                         transport="سيارة (FORD)", price_double="5700")
_qd = build_quotation_data(PassportData(full_name_ar="خالد", room_type="ثنائي"),
                           trip=_tripq, number="MA-Q0001")
assert _qd["stays"] and _qd["flights"] and _qd["guests"], "بيانات غير مكتملة"
# المدينة/الليالي/الفندق/النوع/العدد/الإطلالة/الوجبات + تاريخي الدخول/المغادرة
assert len(_qd["stays"][0]) == 9
assert _qd["gm_name"] == "" and _qd["gm_title"] == ""   # الاسم/الصفة فارغان
# مستند مع إخفاء بنود (طيران + تكلفة)
_qp0 = WORK / "quote_hidden.pdf"
export_umrah_quotation_pdf(PassportData(), _qp0,
                           data={**_qd, "show_flights": False,
                                 "show_costs": False})
assert _qp0.read_bytes()[:5] == b"%PDF-"
# عرض سعر بالإنجليزية (قيَم مترجمة)
_qen = build_quotation_data(PassportData(full_name_ar="خالد", room_type="ثنائي"),
                            trip=_tripq, number="MA-Q0002", lang="en")
assert _qen["lang"] == "en" and _qen["stays"][0][0] == "Madinah"
assert _qen["stays"][0][3] == "Double" and _qen["flight_class"] == "Economy"
_qpen = WORK / "quote_en.pdf"
export_umrah_quotation_pdf(PassportData(), _qpen, data=_qen)
assert _qpen.read_bytes()[:5] == b"%PDF-" and _qpen.stat().st_size > 3000
# تحويل عرض محفوظ عربي ← إنجليزي ← عربي يترجم المحتوى (ذهاباً وإياباً)
from hajj_app.pdf_io import translate_quotation_data as _tq
_qar = build_quotation_data(PassportData(full_name_ar="خالد", room_type="ثنائي"),
                            trip=_tripq, number="MA-Q0003", lang="ar")
_qar["flights"] = [["2026-09-04", "الاتحاد", "14:25", "دبي", "16:15", "جدة"]]
_qar["addressed_title"] = "السيدة"
_qar["note"] = "جميع الحجوزات غير قابلة للإلغاء أو التعديل."
_t_en = _tq(_qar, "en")
assert _t_en["greeting"] == "Greetings," and _t_en["stays"][0][0] == "Madinah"
assert _t_en["flights"][0][1] == "Etihad"        # ترجمة الناقل
assert _t_en["addressed_title"] == "Mrs."        # اللقب (السيدة → Mrs.)
assert _t_en["note"] == "All bookings are non-refundable and non-changeable."
_t_ar = _tq(_t_en, "ar")
assert _t_ar["stays"][0][0] == "المدينة المنوّرة"      # عاد للعربية
assert _t_ar["stays"][0][3] == "ثنائي"
# عرض السعر يتّسع في صفحة واحدة مهما طال (بنود قطار متعددة + مسارات)
_long = dict(_qar)
_long["trains"] = [["2", "سياحية", "المدينة", "مكة", "2026-09-07", "14:30",
                    "16:30"]] * 6
_lp = WORK / "quote_long.pdf"
export_umrah_quotation_pdf(PassportData(), _lp, data=_long)
import fitz as _f
assert _f.open(str(_lp)).page_count == 1, "العرض تجاوز صفحة واحدة"
_qp = WORK / "quote.pdf"
export_umrah_quotation_pdf(PassportData(full_name_ar="خالد"), _qp, data=_qd)
assert _qp.read_bytes()[:5] == b"%PDF-" and _qp.stat().st_size > 3000
# محرّر عرض السعر: قوائم منسدلة، ضيوف، قطار/تأشيرات، تقويم منبثق، معاينة
rq = tk.Tk(); rq.withdraw()
appq = _ug.UmrahApp(rq, session=None)
appq.trips.append(_tripq)
appq.records.append(PassportData(full_name_ar="خالد", trip="Q1",
                                 room_type="ثنائي"))
wq = _ug.TripPilgrimsWindow(appq, _tripq)
wq.tree.selection_set("0")
wq.do_quotation()
_qe = [w for w in wq.winfo_children()
       if isinstance(w, _ug.QuotationEditorDialog)][-1]
assert _qe._number.startswith("MA-Q") and _qe._stay_rows and _qe._flight_rows
# تبديل اللغة إلى الإنجليزية ثم العودة للعربية (يحافظ على الرقم)
_num = _qe._number
_qe._lang_var.set("English"); _qe._on_lang_change()
_qe = [w for w in wq.winfo_children()
       if isinstance(w, _ug.QuotationEditorDialog)][-1]
assert _qe._lang == "en" and _qe._number == _num
_qe._lang_var.set("عربي"); _qe._on_lang_change()
_qe = [w for w in wq.winfo_children()
       if isinstance(w, _ug.QuotationEditorDialog)][-1]
assert _qe._lang == "ar" and _qe._number == _num
assert _qe._stay_rows and _qe._flight_rows
assert _qe._guests and _qe._line_rows       # ضيوف افتراضيون وبنود تنقّل
assert _qe._d.get() and _qe._pf.get()        # حقول التاريخ (تقويم منبثق)
# إضافة ضيوف عبر عناصر التحكّم
_ng = len(_qe._guests)
_qe._g_count.set("1"); _qe._g_type.set("أطفال"); _qe._add_guest_ctl()
assert len(_qe._guests) == _ng + 1
_nf = len(_qe._flight_rows)
_qe._add_flight_row(["2026-09-11", "فلاي دبي", "17:45", "جدة", "21:45", "دبي"])
assert len(_qe._flight_rows) == _nf + 1
# قراءة الطيران من نصّ حجز أماديوس
_amz = umrah.parse_amadeus_flights(
    "1 EY 611 M 04AUG 2 AUHJED DK1 1405 1610\n"
    "2 EY 632 Q 10AUG 1 MEDAUH DK1 1525 1855", year=2026)
assert _amz[0] == ["2026-08-04", "الاتحاد", "14:05", "أبوظبي", "16:10", "جدة"]
assert _amz[1] == ["2026-08-10", "الاتحاد", "15:25", "المدينة", "18:55",
                   "أبوظبي"]
# صيغة OCR مُدمجة بلا مسافات ومع خلط O←0 (كما في لقطات الأماديوس الفعلية)
_amz2 = umrah.parse_amadeus_flights(
    "EY611MO4AUG2AUHJEDDK114051610QO4AUGE06320M\nSEERTSVC\n"
    "EY632Q10AUG1MEDAUHDK11525185510AUGE9321M", year=2026)
assert len(_amz2) == 2, _amz2
assert _amz2[0] == ["2026-08-04", "الاتحاد", "14:05", "أبوظبي", "16:10", "جدة"]
assert _amz2[1] == ["2026-08-10", "الاتحاد", "15:25", "المدينة", "18:55",
                    "أبوظبي"]
# محاكاة زرّ أماديوس في المحرّر (تصحيح OCR وحوار الملفّ)
import hajj_app.ocr as _ocrmod
_ocrmod.read_amadeus_text = lambda p: (
    "1 EY 611 M 04AUG 2 AUHJED DK1 1405 1610\n"
    "2 EY 632 Q 10AUG 1 MEDAUH DK1 1525 1855")
_ug.filedialog.askopenfilename = lambda **k: "dummy.png"
_ug.messagebox.showinfo = lambda *a, **k: None
_qe._amadeus_file()
assert len(_qe._flight_rows) == 2
# سحب وإفلات صورة الأماديوس (لا يتعطّل) — حدث فارغ لا يفعل شيئاً
class _DropEv:
    data = ""
_qe._on_drop_amadeus(_DropEv())
# بنود العرض قابلة للإظهار/الإخفاء
assert {"show_stays", "show_flights", "show_transport", "show_costs"} <= \
    set(_qe._collect())
_qe._show["show_costs"].set(False)
assert _qe._collect()["show_costs"] is False
_qe._show["show_costs"].set(True)
# التوقيع: خانة قابلة للتعديل بهاتف، وخانة المكتب ثابتة
assert "office_name" not in _qe._fields and "gm_phone" in _qe._fields
_qe._fields["gm_phone"].set("0501234567")
_qe._add_line_row(["2026-09-07", "محطة قطار مكة", "فندق مكة"])
# قطار الحرمين: بنود متعددة (تذاكر/درجة/من/إلى/تاريخ/إقلاع/وصول)
_qe._add_train_row(["2", "سياحية", "المدينة", "مكة", "2026-09-07", "14:30",
                    "16:30"])
_qe._add_train_row(["2", "سياحية", "مكة", "المدينة", "2026-09-10", "09:00",
                    "11:00"])
_qe._stay_rows[0][2].set("2026-09-04")         # تاريخ الإقامة يدوياً (من)
_qe._stay_rows[0][3].set("2026-09-07")         # (إلى)
_qe._visas_on.set(True)                        # بند التأشيرات (عدد + نوع)
_qe._visa_count.set("2"); _qe._visa_type.set("عمرة")
_qe._addr_on.set(True)
_qe._addr.set("خالد الشامسي")                  # توجيه باسم الضيف
_qe._vl_on.set(True); _qe._vl.set("2026-07-29"); _qe._vl_time.set("17:00")
_qe._note.insert("1.0", "الأسعار قابلة للتغيير حسب التوفّر.")
_qe._stay_rows[0][1][5].set("مطلّة كعبة")
# التكلفة تُحسب تلقائياً: 2×5700 = 11,400
assert _qe._total_var.get().replace(",", "") == "11400", _qe._total_var.get()
_qe._price_rows[0][1][2].set("3")            # العدد ⇒ 3
assert _qe._total_var.get().replace(",", "") == "17100"
_qe._add_price_row(["أطفال", "ثنائي", "1", "2000"])
assert _qe._total_var.get().replace(",", "") == "19100"
_qdc = _qe._collect()
assert _qdc["number"] == _qe._number and _qdc["flights"] and _qdc["stays"]
assert _qdc["guests"] == [["2", "كبار"], ["1", "أطفال"]]
assert len(_qdc["trains"]) == 2 and _qdc["visas"]   # بندا قطار
assert _qdc["trains"][0][4] == "2026-09-07" and _qdc["trains"][0][5] == "14:30"
assert _qdc["visas"] == "عدد (2) تأشيرة عمرة"        # التأشيرات من العدد والنوع
assert _qdc["addressed_to"] == "خالد الشامسي"     # توجيه باسم الضيف
assert _qdc["validity_time"] == "17:00" and _qdc["note"]   # وقت الصلاحية + ملاحظة
assert _qdc["car_type"] and _qdc["car_model"] and _qdc["car_count"]
assert _qdc["stays"][0][5] == "مطلّة كعبة"
assert len(_qdc["stays"][0]) == 9                         # يشمل تاريخي الدخول/المغادرة
assert _qdc["stays"][0][7] == "2026-09-04"                # تاريخ الإقامة يدوي
# حذف أحد بندي القطار يُبقي بنداً واحداً
_qe._del_row(_qe._train_rows, _qe._train_rows[-1])
assert len(_qe._collect()["trains"]) == 1
assert len(_qdc["pricing"]) == 2 and _qdc["currency"]
# التحقّق من الحساب عبر الدالّة المشتركة
from hajj_app.pdf_io import quotation_pricing as _qpr
assert _qpr(_qdc["pricing"])[1] == 19100.0
# حفظ العرض في «عروض الأسعار» ثم قراءته
_ug.messagebox.showinfo = lambda *a, **k: None
_qe._save()
_saved = umrah.load_quotes(appq._settings, "Q1")
assert len(_saved) == 1 and _saved[0]["number"] == _qe._number
# نافذة عروض الأسعار تعرض العرض المحفوظ
wq.do_quotes_list()
_qlw = [w for w in wq.winfo_children()
        if isinstance(w, _ug.QuotesListWindow)][-1]
assert len(_qlw.tree.get_children()) == 1
umrah.delete_quote(appq._settings, "Q1", _qe._number)
assert umrah.load_quotes(appq._settings, "Q1") == []
_qlw.destroy()
# المعاينة تحفظ العرض تلقائياً في «عروض الأسعار»
_qe._preview()
assert (WORK / "sel.pdf").read_bytes()[:5] == b"%PDF-"
assert len(umrah.load_quotes(appq._settings, "Q1")) == 1   # حُفظ تلقائياً
_qe.destroy()
# عرض سعر يدوي خارج البرامج
appq.new_manual_quotation()
_qm = [w for w in rq.winfo_children()
       if isinstance(w, _ug.QuotationEditorDialog)][-1]
assert _qm.trip is None and _qm._number.startswith("MA-Q")
_qm._add_stay_row(["مكة المكرّمة", "4", "فندق النور", "ثنائي", "1",
                   "مطلّة كعبة", "إفطار"])
_qm._add_guest([" 2", "كبار"])
_qm._preview()
assert (WORK / "sel.pdf").read_bytes()[:5] == b"%PDF-"
_qm.destroy()
rq.destroy()
print("  OK: عرض السعر (Quotation) — مستند ومحرّر ونسخة يدوية")

# === مسعّر المجموعات ===
from hajj_app.pdf_io import export_group_pricing_pdf
_gd = dict(makkah_rate="1426", makkah_nights="3", transport="50",
           ticket="1265", water="46", gifts="150", admin="100", profit="200",
           currency="درهم")
_grows = umrah.group_pricing(_gd)
_d2 = next(r for r in _grows if r["type"] == "ثنائي")
assert _d2["net"] == 3750.0 and _d2["selling"] == 3950.0     # مطابقة الجدول
_s1 = next(r for r in _grows if r["type"] == "مفرد")
assert _s1["net"] == 5889.0                                  # المفرد = الغرفة كاملة
# نسبة ربح مئوية (10% من الصافي)
_gpct = umrah.group_pricing({**_gd, "profit": "", "profit_pct": "10"})
_dp = next(r for r in _gpct if r["type"] == "ثنائي")
assert _dp["margin"] == 375.0 and _dp["selling"] == 4125.0
# مبلغ ربح لكل نوع غرفة
_gper = umrah.group_pricing({k: v for k, v in _gd.items() if k != "profit"}
                            | {"profit_single": "500", "profit_double": "300"})
assert next(r for r in _gper if r["type"] == "مفرد")["margin"] == 500.0
assert next(r for r in _gper if r["type"] == "ثنائي")["margin"] == 300.0
# اختيار أنواع الغرف المطلوب تسعيرها فقط
_gsel = umrah.group_pricing({**_gd, "room_types": ["ثنائي", "ثلاثي"]})
assert [r["type"] for r in _gsel] == ["ثنائي", "ثلاثي"]
assert umrah.group_pricing({**_gd, "room_types": []}) == _grows   # فارغ = الكل
# حذف المدينة أو تضمينها
_gmd = {**_gd, "madinah_rate": "1000", "madinah_nights": "2"}
assert all(r["madinah"] == 0.0
           for r in umrah.group_pricing({**_gmd, "include_madinah": "0"}))
assert next(r for r in umrah.group_pricing({**_gmd, "include_madinah": "1"})
            if r["type"] == "ثنائي")["madinah"] == 1000.0
assert next(r for r in umrah.group_pricing(_gmd)   # غياب المفتاح = مُضمّنة
            if r["type"] == "ثنائي")["madinah"] == 1000.0
_gp = WORK / "group.pdf"
export_group_pricing_pdf(_gd, _gp)
assert _gp.read_bytes()[:5] == b"%PDF-" and _gp.stat().st_size > 3000
export_group_pricing_pdf({**_gd, "room_types": ["ثنائي"]}, WORK / "gsel.pdf")
assert (WORK / "gsel.pdf").read_bytes()[:5] == b"%PDF-"
# نافذة المسعّر مع حساب حيّ
rg = tk.Tk(); rg.withdraw()
appg = _ug.UmrahApp(rg, session=None)
appg.open_group_pricer()
_gw = [w for w in rg.winfo_children()
       if isinstance(w, _ug.GroupPricerWindow)][-1]
_gw._f["title"].set("تسعير مجموعة")             # عنوان التسعير
_gw._f["makkah_rate"].set("1426"); _gw._f["makkah_nights"].set("3")
assert len(_gw._item_rows) == 7                 # بنود افتراضية
# البنود ديناميكية: النقل(0)/المطار(1)/التأشيرة(2)/التذكرة(3)/ماء(4)/هدايا(5)/إداري(6)
_gw._item_rows[0][2].set("50"); _gw._item_rows[3][2].set("1265")
_gw._item_rows[4][2].set("46"); _gw._item_rows[5][2].set("150")
_gw._item_rows[6][2].set("100")
_gw._f["profit"].set("200")
_gw._add_item_row("بند مخصّص", "0")             # إضافة بند
assert len(_gw._item_rows) == 8
_gw._del_row(_gw._item_rows, _gw._item_rows[-1]); _gw._recalc()   # إلغاء بند
assert len(_gw._item_rows) == 7
rg.update()
_vals = {row[0]: row for row in
         (_gw._tree.item(i, "values") for i in _gw._tree.get_children())}
assert _vals["ثنائي"][1].replace(",", "") == "3750"        # الصافية
assert _vals["ثنائي"][4].replace(",", "") == "3950"        # سعر البيع
assert _vals["ثنائي"][3].endswith("%")                     # النسبة تلقائية
# اختيار نوع الغرفة: إلغاء المفرد يخفيه من النتيجة
_gw._type_vars["مفرد"].set(False)
rg.update()
_shown = [_gw._tree.item(i, "values")[0] for i in _gw._tree.get_children()]
assert "مفرد" not in _shown and "ثنائي" in _shown
assert _gw._collect()["room_types"] == ["ثنائي", "ثلاثي", "رباعي", "طفل"]
_gw._type_vars["مفرد"].set(True)
rg.update()
# حذف/تضمين المدينة
_gw._f["madinah_rate"].set("1000"); _gw._f["madinah_nights"].set("2")
_gw._inc_md.set(False); _gw._toggle_madinah(); rg.update()
assert _gw._collect()["include_madinah"] == "0"
assert str(_gw._widgets["madinah_rate"].cget("state")) == "disabled"
_dbl2 = next(r for r in umrah.group_pricing(_gw._collect())
             if r["type"] == "ثنائي")
assert _dbl2["madinah"] == 0.0
_gw._inc_md.set(True); _gw._toggle_madinah(); rg.update()
assert str(_gw._widgets["madinah_rate"].cget("state")) == "normal"
_gw._preview()
assert (WORK / "sel.pdf").read_bytes()[:5] == b"%PDF-"
# حفظ التسعير داخل البرنامج واستعراضه لاحقاً
assert _gw._number.startswith("MA-P")
_gw._save()
_psaved = umrah.load_pricings(appg._settings)
assert len(_psaved) == 1 and _psaved[0]["number"] == _gw._number
_gw.destroy()
appg.open_pricings()
_plw = [w for w in rg.winfo_children()
        if isinstance(w, _ug.PricingsListWindow)][-1]
assert len(_plw.tree.get_children()) == 1
_plw.tree.selection_set("0"); _plw.open_sel()
_ped = [w for w in _plw.winfo_children()
        if isinstance(w, _ug.GroupPricerWindow)][-1]
assert _ped._number == _gw._number and len(_ped._item_rows) == 7
_ped._save()                                    # نفس الرقم، بلا تكرار
assert len(umrah.load_pricings(appg._settings)) == 1
umrah.delete_pricing(appg._settings, _gw._number)
assert umrah.load_pricings(appg._settings) == []
_ped.destroy(); _plw.destroy()
rg.destroy()
print("  OK: مسعّر المجموعات — حساب ومستند وحفظ")

app_mode.set_mode("hajj")
print("\n*** UMRAH TESTS PASSED ***")

"""استيراد وتصدير قوائم الحجاج بصيغة إكسل (.xlsx)."""

from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .fields import (
    DATE_KEYS, FIELDS, MONEY_KEYS, TIME_KEYS, _norm, format_amount, match_column,
    normalize_time, parse_amount, row_dict,
)
from .mrz import PassportData

_HEADER_FILL = PatternFill("solid", fgColor="1F6F4A")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_WARN_FILL = PatternFill("solid", fgColor="FFF3CD")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _to_text(value) -> str:
    """يحوّل قيمة خلية إلى نص، مع تنسيق التواريخ بصيغة موحّدة."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_time_text(value) -> str:
    """يقرأ خلية وقت. إكسل يخزّن الوقت ككائن time أو datetime بتاريخ وهمي."""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float) and 0 <= value < 1:
        # كسر من اليوم (تنسيق إكسل الداخلي للوقت)
        minutes = round(value * 24 * 60)
        return f"{minutes // 60:02d}:{minutes % 60:02d}"
    return normalize_time(_to_text(value))


def _normalize_date(text: str) -> str:
    """يحاول توحيد التاريخ إلى YYYY-MM-DD من الصيغ الشائعة."""
    text = text.strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d",
                "%m/%d/%Y", "%d.%m.%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    # صيغة YYMMDD أو YYYYMMDD بلا فواصل
    if re.fullmatch(r"\d{8}", text):
        try:
            return datetime.strptime(text, "%Y%m%d").date().isoformat()
        except ValueError:
            pass
    return text  # نُبقيه كما هو بدل حذف بيانات المستخدم


def export_excel(records: list[PassportData], path: str | Path) -> Path:
    """يصدّر السجلات إلى ملف إكسل منسّق."""
    path = Path(path)
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "الحجاج"
    ws.sheet_view.rightToLeft = True

    # العنوان الرئيسي
    ws.append([f"كشف الحجاج — {date.today().isoformat()}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(FIELDS))
    title = ws.cell(row=1, column=1)
    title.font = Font(bold=True, size=14, color="1F6F4A")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # رؤوس الأعمدة
    ws.append([f.label for f in FIELDS])
    for col in range(1, len(FIELDS) + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[2].height = 34

    # البيانات
    for idx, rec in enumerate(records, start=1):
        data = row_dict(rec, idx)
        ws.append([data.get(f.key, "") for f in FIELDS])
        row = ws.max_row
        has_warning = bool(data.get("warnings"))
        for col, f in enumerate(FIELDS, start=1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER
            if has_warning:
                cell.fill = _WARN_FILL
            # المبالغ كأرقام حقيقية ليصح الجمع في إكسل
            if f.key in MONEY_KEYS:
                amount = parse_amount(cell.value)
                if amount is not None:
                    cell.value = amount
                    cell.number_format = "#,##0"

    # عرض الأعمدة + تجميد الرؤوس + فلتر
    for i, f in enumerate(FIELDS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = f.width
    # نجمّد المسلسل والأسماء ليبقيا ظاهرين عند التمرير أفقياً
    ws.freeze_panes = "F3"
    if records:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(FIELDS))}{ws.max_row}"

    wb.save(path)
    return path


def export_umrah_excel(records: list[PassportData], path: str | Path, *,
                       program_name: str = "", manager: str = "") -> Path:
    """يصدّر كشف معتمري برنامج عمرة إلى إكسل بمسمّيات العمرة والأعمدة المطلوبة."""
    from .umrah import REPORT_COLUMNS, REPORT_MONEY_KEYS, report_row

    path = Path(path)
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "المعتمرون"
    ws.sheet_view.rightToLeft = True
    ncols = len(REPORT_COLUMNS)

    heading = f"كشف المعتمرين — {program_name}" if program_name else "كشف المعتمرين"
    if manager:
        heading += f" — المسؤول: {manager}"
    ws.append([f"{heading} — {date.today().isoformat()}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title = ws.cell(row=1, column=1)
    title.font = Font(bold=True, size=14, color="1F6F4A")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.append([lbl for _k, lbl in REPORT_COLUMNS])
    for col in range(1, ncols + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[2].height = 32

    for idx, rec in enumerate(records, start=1):
        row = report_row(rec, idx, program_name)
        ws.append([row[k] for k, _l in REPORT_COLUMNS])
        r = ws.max_row
        for col, (key, _lbl) in enumerate(REPORT_COLUMNS, start=1):
            cell = ws.cell(row=r, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER
            if key in REPORT_MONEY_KEYS:
                amount = parse_amount(cell.value)
                if amount is not None:
                    cell.value = amount
                    cell.number_format = "#,##0"

    widths = (10, 26, 12, 15, 16, 13, 18, 20, 12, 15, 12, 14, 12)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C3"
    if records:
        ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{ws.max_row}"

    wb.save(path)
    return path


# الحقول التي تُعرّف الحاج. صف بلا أيٍّ منها ليس سجل حاج مهما امتلأت بقية
# خاناته — وقوائم كثيرة تحوي صفوف ترقيم فارغة أو صفوف مجاميع.
_IDENTITY_KEYS = (
    "full_name_ar", "full_name_en", "given_names_en", "surname_en",
    "passport_number",
)


def _count_usable_rows(ws: Worksheet, header_row: int,
                       mapping: dict[int, str], limit: int = 400) -> int:
    """يعدّ صفوف الورقة التي تحمل هوية فعلية (اسم أو رقم جواز).

    نقيس البيانات لا الرؤوس: قالب فارغ قد يحمل رؤوساً مثالية وصفوفاً خاوية،
    فاختياره على ورقة فيها بيانات حقيقية يضيّع الكشف كله.
    """
    columns = [c for c, key in mapping.items() if key in _IDENTITY_KEYS]
    if not columns:
        return 0
    count = 0
    for offset, row in enumerate(
        ws.iter_rows(min_row=header_row + 2, values_only=True)
    ):
        if offset >= limit:
            break
        if any(c < len(row) and _to_text(row[c]) for c in columns):
            count += 1
    return count


# رؤوس "وقت" مجرّدة لا تحدّد وصولاً من مغادرة بذاتها — نستدلّ من عمود
# التاريخ المجاور (الملاحق أو السابق) على أيّهما تقصد.
_BARE_TIME_NORMS = frozenset(
    _norm(t) for t in ("الوقت", "الساعة", "الزمن", "time", "hour")
)


def _resolve_bare_times(row, mapping: dict[int, str]) -> None:
    """يعيّن أعمدة «الوقت» المجرّدة إلى وقت الوصول/المغادرة حسب جارها.

    ملفات كثيرة تكتب عمود الوقت باسم «الوقت» فقط بجوار «تاريخ الوصول» أو
    «تاريخ المغادرة». نربطه بالتاريخ المجاور له (التالي أولاً ثم السابق).
    """
    used = set(mapping.values())
    for c_idx, cell in enumerate(row):
        if c_idx in mapping or cell is None:
            continue
        if _norm(_to_text(cell)) not in _BARE_TIME_NORMS:
            continue
        for j in (c_idx + 1, c_idx - 1):          # التاريخ يلي الوقت عادةً
            date_key = mapping.get(j) if 0 <= j < len(row) else None
            if date_key == "arrival_date" and "arrival_time" not in used:
                mapping[c_idx] = "arrival_time"
                used.add("arrival_time")
                break
            if date_key == "departure_date" and "departure_time" not in used:
                mapping[c_idx] = "departure_time"
                used.add("departure_time")
                break


def _detect_header(ws: Worksheet) -> tuple[int, dict[int, str]]:
    """يجد صف الرؤوس في ورقة ويعيد (رقمه، خريطة العمود -> الحقل).

    لا نفترض أنه الصف الأول: كثير من الكشوف تبدأ بعنوان تزييني أو صف فارغ.
    """
    best_row, best_map = -1, {}
    for r_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        mapping: dict[int, str] = {}
        for c_idx, cell in enumerate(row):
            if cell is None:
                continue
            key = match_column(_to_text(cell))
            if key and key not in mapping.values():
                mapping[c_idx] = key
        _resolve_bare_times(row, mapping)
        if len(mapping) > len(best_map):
            best_row, best_map = r_idx, mapping
    return best_row, best_map


_ROOM_FILL = PatternFill("solid", fgColor="3A342B")
_ROOM_FONT = Font(bold=True, color="FFFFFF")


def export_grouped_excel(records, path: str | Path, *, title: str = "كشف التسكين") -> Path:
    """يصدّر كشف التسكين إلى إكسل بنفس أعمدة الطباعة، مجموعاً بالغرف.

    كل غرفة كتلة يسبقها سطر عنوان داكن يمتد عبر الأعمدة، فتظهر منفصلة عن
    التي تليها — مطابقة لتخطيط الطباعة. من لا غرفة له في كتلة "بدون غرفة".
    """
    from .fields import PDF_FIELDS
    from .rooming import common_room_type, group_records_by_room

    path = Path(path)
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "التسكين"
    ws.sheet_view.rightToLeft = True

    cols = list(PDF_FIELDS)
    ncols = len(cols)
    ws.append([f.label for f in cols])
    for col, f in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = f.width
    ws.row_dimensions[1].height = 30

    def add_room_header(text: str) -> None:
        ws.append([text] + [""] * (ncols - 1))
        row = ws.max_row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1)
        cell.fill = _ROOM_FILL
        cell.font = _ROOM_FONT
        cell.alignment = Alignment(horizontal="right", vertical="center")

    def add_occupant(rec, serial: int) -> None:
        data = row_dict(rec, serial)
        ws.append([data.get(f.key, "") for f in cols])
        row = ws.max_row
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

    rooms, unplaced = group_records_by_room(records)
    show_hotel = len({hotel for hotel, *_ in rooms if hotel}) > 1
    serial = 0
    for hotel, cap, number, occ in rooms:
        rtype = common_room_type(occ) or "غرفة"
        label = (f"غرفة {number} — {rtype} ({len(occ)}/{cap})" if number
                 else f"{rtype} ({len(occ)})")
        if show_hotel and hotel:
            label = f"{hotel} — {label}"
        add_room_header(label)
        for rec in occ:
            serial += 1
            add_occupant(rec, serial)
    if unplaced:
        add_room_header(f"بدون غرفة ({len(unplaced)})")
        for rec in unplaced:
            serial += 1
            add_occupant(rec, serial)

    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def import_excel(path: str | Path) -> tuple[list[PassportData], list[str]]:
    """يستورد سجلات من ملف إكسل.

    يفحص **كل أوراق الملف** ويختار أنسبها، لأن الورقة النشطة كثيراً ما
    تكون ملخّصاً أو كشف غرف بينما البيانات الكاملة في ورقة أخرى.

    يبحث عن صف الرؤوس تلقائياً (قد لا يكون الصف الأول بسبب عناوين تزيينية)
    ويطابق الأعمدة بالمرادفات، فيعمل مع ملفات من جهات مختلفة.

    يعيد: (السجلات، رسائل تنبيه عن الأعمدة غير المعروفة)
    """
    path = Path(path)
    wb = load_workbook(path, data_only=True)

    # أنسب ورقة = أكثرها صفوفاً تحمل هوية فعلية، ثم أكثرها أعمدة معروفة.
    # الترتيب مقصود: ورقة بأعمدة أقل وبيانات حقيقية خير من قالب مثالي فارغ.
    candidates = []
    for sheet in wb.worksheets:
        header_row, mapping = _detect_header(sheet)
        if mapping:
            usable = _count_usable_rows(sheet, header_row, mapping)
            candidates.append((usable, len(mapping), sheet, header_row, mapping))

    if not candidates:
        return [], [
            "لم يتم التعرّف على أي عمود في الملف.\n"
            "تأكد أن الصف الأول يحتوي عناوين مثل: الاسم، رقم الجواز، الجنسية، تاريخ الميلاد."
        ]

    usable_count, _matched, ws, best_row, best_map = max(
        candidates, key=lambda c: (c[0], c[1])
    )
    if usable_count == 0:
        return [], [
            "تم التعرّف على الأعمدة لكن لا يوجد صف يحمل اسماً أو رقم جواز.\n"
            "تأكد أن الملف يحتوي بيانات وليس قالباً فارغاً."
        ]

    notes: list[str] = []
    if len(wb.worksheets) > 1:
        notes.append(f"قُرئت الورقة: «{ws.title}» (من {len(wb.worksheets)} أوراق)")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["الملف فارغ"]
    unknown = [
        _to_text(c) for i, c in enumerate(rows[best_row])
        if c is not None and i not in best_map and _to_text(c)
    ]
    if unknown:
        notes.append("أعمدة لم يتم التعرّف عليها (تم تجاهلها): " + "، ".join(unknown))

    records: list[PassportData] = []
    skipped = 0

    for row in rows[best_row + 1:]:
        values = {}
        for c_idx, key in best_map.items():
            if c_idx >= len(row):
                continue
            text = _to_text(row[c_idx])
            if key in DATE_KEYS:
                text = _normalize_date(text)
            elif key in TIME_KEYS:
                text = _to_time_text(row[c_idx])
            elif key in MONEY_KEYS:
                amount = parse_amount(text)
                text = format_amount(amount) if amount is not None else text
            values[key] = text

        # صف بلا اسم ولا رقم جواز ليس سجل حاج: قد يكون صف ترقيم في قالب
        # فارغ، أو صف مجاميع، أو فاصلاً. استيراده يملأ الكشف بسجلات خاوية.
        if not any(values.get(key, "") for key in _IDENTITY_KEYS):
            skipped += 1
            continue

        rec = PassportData(source_file=path.name)
        for key, text in values.items():
            # المسلسل والمتبقي محسوبان، فلا نستوردهما
            if key in ("serial", "remaining_amount"):
                continue
            if hasattr(rec, key):
                setattr(rec, key, text)

        # نستكمل ما يمكن اشتقاقه
        if not rec.full_name_en and (rec.given_names_en or rec.surname_en):
            rec.full_name_en = " ".join(
                p for p in (rec.given_names_en, rec.surname_en) if p
            )
        if not rec.nationality_ar and rec.nationality:
            from .mrz import COUNTRY_AR
            rec.nationality_ar = COUNTRY_AR.get(rec.nationality, rec.nationality)

        records.append(rec)

    if skipped:
        notes.append(f"تُخطّي {skipped} صفاً بلا اسم ولا رقم جواز")
    if not records:
        notes.append("تم التعرّف على الأعمدة لكن لا توجد صفوف بيانات.")

    return records, notes

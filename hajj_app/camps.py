"""كشف تسكين المخيمات (منى وعرفة): توزيع الحجّاج على الخيام.

القواعد المعتمدة:
- **الرجال في خيام والنساء في خيام منفصلة** — لا اختلاط في الخيمة الواحدة.
- أفراد العائلة الواحدة (نفس رقم العائلة) معاً، وسكّان الغرفة الواحدة معاً؛
  فمن جمعتهم غرفة أو عائلة لا يُفرّقون إلا للفصل بين الجنسين.
- المستخدم يحدّد: **عدد الأشخاص في الخيمة**، و**القطاع**، و**رقم الخيمة الأول**.
- **التصنيف** (رجال/نساء) يُستنبط من جنس الحاج ويظهر لكل خيمة.
- التوزيع يُبنى عند الطلب ولا يُحفظ في بيانات الحاج.

مخيّم منى ومخيّم عرفة يستعملان الخوارزمية نفسها؛ اسم المخيّم مجرّد عنوان.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .mrz import PassportData
from .rooming import Occupant, room_capacity, room_number_in_type

CAMP_MINA = "منى"
CAMP_ARAFAT = "عرفة"
CAMPS: tuple[str, ...] = (CAMP_MINA, CAMP_ARAFAT)

MEN = "رجال"
WOMEN = "نساء"
UNKNOWN = "غير محدد"
_CLASS_ORDER: tuple[str, ...] = (MEN, WOMEN, UNKNOWN)

_DEFAULT_CAPACITY = 40


def classification(sex: str) -> str:
    """تصنيف الحاج للفصل في الخيام: رجال/نساء، أو «غير محدد» إن غاب الجنس."""
    s = str(sex or "").strip()
    if s.startswith("ذكر") or s.upper().startswith("M"):
        return MEN
    if s.startswith("أنثى") or s.startswith("انثى") or s.upper().startswith("F"):
        return WOMEN
    return UNKNOWN


def _room_key(rec: PassportData):
    """مفتاح الغرفة (فندق، سعة، رقم) إن كان للحاج غرفة فعلية، وإلا None."""
    rtype = str(rec.room_type or "")
    number = str(rec.room_number or "").strip() or room_number_in_type(rtype)
    if not number:
        return None
    return ("r", str(rec.hotel or "").strip(), room_capacity(rtype), number)


def _cluster_records(records: list[PassportData]) -> list[list[int]]:
    """يجمع الحجّاج في عناقيد لا تُفرَّق: من جمعتهم عائلة أو غرفة في عنقود واحد.

    يعيد قائمة عناقيد (فهارس السجلات)، محافظاً على ترتيب أول ظهور.
    """
    n = len(records)
    parent = list(range(n))

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[max(ra, rb)] = min(ra, rb)

    groups: dict[tuple, list[int]] = {}
    for i, rec in enumerate(records):
        fam = str(rec.family_number or "").strip()
        if fam:
            groups.setdefault(("f", fam), []).append(i)
        room = _room_key(rec)
        if room is not None:
            groups.setdefault(room, []).append(i)

    for members in groups.values():
        first = members[0]
        for other in members[1:]:
            union(first, other)

    clusters: dict[int, list[int]] = {}
    order: list[int] = []
    for i in range(n):
        root = find(i)
        if root not in clusters:
            clusters[root] = []
            order.append(root)
        clusters[root].append(i)
    return [clusters[root] for root in order]


@dataclass
class Tent:
    """خيمة في كشف المخيمات."""
    camp: str
    sector: str
    number: str
    classification: str
    capacity: int
    occupants: list[Occupant] = field(default_factory=list)

    @property
    def count(self) -> int:
        return len(self.occupants)

    @property
    def families(self) -> list[str]:
        seen: list[str] = []
        for occ in self.occupants:
            fam = occ.family or "—"
            if fam not in seen:
                seen.append(fam)
        return seen

    def warnings(self) -> list[str]:
        issues = []
        if self.count > self.capacity:
            issues.append(f"عدد الأشخاص {self.count} يتجاوز سعة الخيمة {self.capacity}")
        return issues


@dataclass
class CampPlan:
    """نتيجة توزيع مخيّم: الخيام والتنبيهات."""
    camp: str
    sector: str
    capacity: int
    tents: list[Tent] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    @property
    def total(self) -> int:
        return sum(t.count for t in self.tents)


def _pack(units: list[list[Occupant]], capacity: int,
          notes: list[str], cls: str) -> list[list[Occupant]]:
    """يوزّع الوحدات (عائلة/غرفة) على خيام بحجم السعة، مع بقاء كل وحدة معاً.

    ملء أوّل ملائم (first-fit): تُوضع الوحدة في أول خيمة يتّسع باقيها لها،
    وإلا فُتحت خيمة جديدة. الوحدة الأكبر من السعة تُقسَّم مع تنبيه.
    """
    tents: list[list[Occupant]] = []
    remaining: list[int] = []
    for unit in units:
        if len(unit) > capacity:
            notes.append(
                f"وحدة ({cls}) فيها {len(unit)} شخصاً تتجاوز سعة الخيمة "
                f"{capacity} — قُسّمت على أكثر من خيمة"
            )
            for start in range(0, len(unit), capacity):
                chunk = unit[start:start + capacity]
                tents.append(list(chunk))
                remaining.append(capacity - len(chunk))
            continue
        placed = False
        for idx in range(len(tents)):
            if remaining[idx] >= len(unit):
                tents[idx].extend(unit)
                remaining[idx] -= len(unit)
                placed = True
                break
        if not placed:
            tents.append(list(unit))
            remaining.append(capacity - len(unit))
    return tents


def build_camp_plan(
    records: list[PassportData],
    camp: str = CAMP_MINA,
    *,
    capacity: int = _DEFAULT_CAPACITY,
    sector: str = "",
    start_number: int = 1,
    only: str = "",
) -> CampPlan:
    """يبني خطة تسكين مخيّم: يوزّع الحجّاج على خيام مفصولة بالجنس.

    كل عنقود (عائلة/سكّان غرفة) يُقسَّم بالتصنيف فيبقى رجاله معاً ونساؤه معاً،
    ثم تُملأ الخيام بحجم السعة. الخيام تُرقّم تسلسلياً من `start_number`:
    خيام الرجال أولاً ثم النساء ثم غير المحدّدين.

    `only`: إن حُدّد تصنيف (رجال/نساء/غير محدد) يقتصر الكشف عليه فقط،
    وتُرقّم خيامه من `start_number` — لطباعة كشف الرجال أو النساء وحده.
    """
    try:
        capacity = max(1, int(capacity))
    except (TypeError, ValueError):
        capacity = _DEFAULT_CAPACITY
    try:
        number = int(str(start_number).strip() or 1)
    except (TypeError, ValueError):
        number = 1
    sector = str(sector or "").strip()
    only = str(only or "").strip()
    classes = (only,) if only in _CLASS_ORDER else _CLASS_ORDER

    occ_by_index = {i: Occupant(i + 1, rec) for i, rec in enumerate(records)}
    clusters = _cluster_records(records)

    # وحدات لكل تصنيف: العنقود يُقسَّم بالجنس مع بقاء أفراد كل جنس معاً
    units: dict[str, list[list[Occupant]]] = {c: [] for c in _CLASS_ORDER}
    for cluster in clusters:
        by_class: dict[str, list[Occupant]] = {}
        for i in cluster:
            cls = classification(records[i].sex)
            by_class.setdefault(cls, []).append(occ_by_index[i])
        for cls, occs in by_class.items():
            units[cls].append(occs)

    notes: list[str] = []
    tents: list[Tent] = []
    for cls in classes:
        if not units[cls]:
            continue
        for group in _pack(units[cls], capacity, notes, cls):
            tents.append(Tent(
                camp=camp, sector=sector, number=str(number),
                classification=cls, capacity=capacity, occupants=group,
            ))
            number += 1

    if UNKNOWN in classes and units[UNKNOWN]:
        count = sum(len(u) for u in units[UNKNOWN])
        notes.append(
            f"{count} حاجاً بلا جنس محدّد — أُدرجوا في خيام «غير محدد»؛ "
            "حدّد جنسهم ليُفصلوا بشكل صحيح"
        )
    return CampPlan(camp=camp, sector=sector, capacity=capacity,
                    tents=tents, notes=notes)


def _room_of(rec: PassportData) -> str:
    return (str(rec.room_number or "").strip()
            or room_number_in_type(str(rec.room_type or "")))


def camp_rows(plan: CampPlan) -> list[dict]:
    """يحوّل الخطة إلى صفوف عرض مسطّحة (خيمة ثم سكانها)، للجدول والتصدير."""
    rows: list[dict] = []
    serial = 0
    for tent in plan.tents:
        for position, occ in enumerate(tent.occupants, start=1):
            serial += 1
            rows.append({
                "camp": tent.camp,
                "sector": tent.sector,
                "tent": tent.number,
                "classification": tent.classification,
                "position": position,
                "serial": serial,
                "name": occ.name,
                "family_number": occ.family,
                "hotel": str(occ.record.hotel or "").strip(),
                "room": _room_of(occ.record),
                "sex": occ.sex,
                "phone": str(occ.record.phone or "").strip(),
                "warnings": tent.warnings(),
            })
    return rows


# ------------------------------------------------------------------ التصدير
# أعمدة الكشف (تُرتّب من اليمين لليسار في العرض العربي)
CAMP_COLUMNS: tuple[str, ...] = (
    "م", "اسم الحاج", "رقم العائلة", "الفندق", "الغرفة", "الجنس", "الهاتف",
)
_CAMP_WIDTHS = (5, 28, 10, 18, 8, 8, 14)

_HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# لون رأس الخيمة حسب التصنيف
CLASS_FILL_HEX = {MEN: "2F6F76", WOMEN: "8A4B52", UNKNOWN: "3A342B"}
_CLASS_FONT = Font(bold=True, color="FFFFFF")


def tent_label(tent: Tent) -> str:
    """عنوان رأس الخيمة: «خيمة 3 — رجال — قطاع ب (38/40)»."""
    label = f"خيمة {tent.number} — {tent.classification}"
    if tent.sector:
        label += f" — قطاع {tent.sector}"
    return f"{label}  ({tent.count}/{tent.capacity})"


def export_camp_excel(plan: CampPlan, path: str | Path,
                      *, title: str = "كشف تسكين المخيمات") -> Path:
    """يصدّر خطة المخيّم إلى إكسل — مجموعاً بالخيام، كل خيمة يسبقها سطر عنوان."""
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = plan.camp or "المخيمات"
    ws.sheet_view.rightToLeft = True
    ncols = len(CAMP_COLUMNS)

    # عنوان علوي يمتد عبر الأعمدة
    full_title = f"{title} — مخيّم {plan.camp}" if plan.camp else title
    ws.append([full_title] + [""] * (ncols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1)
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # صف الرؤوس
    ws.append(list(CAMP_COLUMNS))
    for col, width in enumerate(_CAMP_WIDTHS, start=1):
        cell = ws.cell(row=2, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 26

    def add_tent_header(tent: Tent) -> None:
        ws.append([tent_label(tent)] + [""] * (ncols - 1))
        row = ws.max_row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1)
        cell.fill = PatternFill("solid", fgColor=CLASS_FILL_HEX.get(
            tent.classification, CLASS_FILL_HEX[UNKNOWN]))
        cell.font = _CLASS_FONT
        cell.alignment = Alignment(horizontal="right", vertical="center")

    serial = 0
    for tent in plan.tents:
        add_tent_header(tent)
        for occ in tent.occupants:
            serial += 1
            ws.append([
                serial, occ.name, occ.family,
                str(occ.record.hotel or "").strip(), _room_of(occ.record),
                occ.sex, str(occ.record.phone or "").strip(),
            ])
            row = ws.max_row
            for col in range(1, ncols + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = _BORDER

    ws.freeze_panes = "A3"
    wb.save(path)
    return path


# ---- كشف كل خيمة على حدة (أعمدة مبسّطة: تسلسل، اسم، قطاع، خيمة، تصنيف، حملة)
TENT_SHEET_COLUMNS: tuple[str, ...] = (
    "التسلسل", "اسم الحاج", "القطاع", "خيمة رقم", "التصنيف", "اسم الحملة",
)
_TENT_WIDTHS = (8, 32, 10, 10, 10, 26)
_INVALID_SHEET = re.compile(r"[:\\/?*\[\]]")


def _sheet_name(base: str, used: set[str], fallback: str) -> str:
    """اسم ورقة إكسل صالح (≤31 حرفاً، بلا رموز ممنوعة، فريد)."""
    name = _INVALID_SHEET.sub("-", str(base)).strip()[:28] or fallback
    candidate, n = name, 2
    while candidate in used:
        candidate = f"{name} ({n})"
        n += 1
    used.add(candidate)
    return candidate


def export_tents_excel(plan: CampPlan, path: str | Path,
                       *, campaign: str = "") -> Path:
    """يصدّر كل خيمة في **ورقة مستقلة**، بالأعمدة المبسّطة فقط.

    الأعمدة: التسلسل • اسم الحاج • القطاع • خيمة رقم • التصنيف • اسم الحملة.
    """
    path = Path(path)
    campaign = str(campaign or "").strip()
    ncols = len(TENT_SHEET_COLUMNS)
    wb = Workbook()
    wb.remove(wb.active)                     # نبدأ بلا ورقة افتراضية
    used: set[str] = set()

    for index, tent in enumerate(plan.tents, start=1):
        ws = wb.create_sheet(_sheet_name(f"خيمة {tent.number}", used, f"خيمة {index}"))
        ws.sheet_view.rightToLeft = True
        ws.append(list(TENT_SHEET_COLUMNS))
        for col, width in enumerate(_TENT_WIDTHS, start=1):
            cell = ws.cell(row=1, column=col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _BORDER
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 26
        for position, occ in enumerate(tent.occupants, start=1):
            ws.append([position, occ.name, tent.sector, tent.number,
                       tent.classification, campaign])
            row = ws.max_row
            for col in range(1, ncols + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = _BORDER
        ws.freeze_panes = "A2"

    if not wb.sheetnames:                     # لا خيام — نترك ورقة فارغة
        wb.create_sheet("لا خيام")
    wb.save(path)
    return path

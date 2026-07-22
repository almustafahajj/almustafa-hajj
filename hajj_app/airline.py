"""كشف الطيران بالإنجليزية (LTR) وإدخالات أماديوس.

تصدير قائمة ركّاب بالأعمدة التي تطلبها شركات الطيران، بالإنجليزية ومن
اليسار لليمين، مع توليد أوامر إدخال أماديوس (اسم + وثيقة السفر DOCS)
جاهزة للنسخ.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .mrz import COUNTRY_AR, PassportData

# أعمدة كشف الطيران بالترتيب المطلوب (من اليسار لليمين)
AIRLINE_COLUMNS = (
    "#", "Last Name", "First Name", "Passport No.", "Expiry", "DOB",
    "Gender", "Nationality", "Class (J or Y)", "Family No.", "PNR",
)

_AR_TO_CODE = {name: code for code, name in COUNTRY_AR.items()}
_MONTHS = ("JAN", "FEB", "MAR", "APR", "MAY", "JUN",
           "JUL", "AUG", "SEP", "OCT", "NOV", "DEC")


def _clean_latin(text: str) -> str:
    """يبقي الحروف اللاتينية والفراغات، بأحرف كبيرة (لأماديوس والطيران)."""
    text = re.sub(r"[^A-Za-z ]+", " ", str(text or "")).upper()
    return re.sub(r"\s+", " ", text).strip()


def split_name(rec: PassportData) -> tuple[str, str]:
    """يعيد (اللقب، الاسم الأول) بالإنجليزية.

    يفضّل ما جاء من MRZ؛ وإن غاب يقسّم الاسم الكامل — آخر كلمة لقباً
    (اسم القبيلة/العائلة في التسمية العربية) والباقي اسماً أول.
    """
    last = _clean_latin(rec.surname_en)
    first = _clean_latin(rec.given_names_en)
    if last or first:
        return last, first
    parts = _clean_latin(rec.full_name_en).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[-1], " ".join(parts[:-1])


def gender_code(rec: PassportData) -> str:
    """M أو F من حقل الجنس، أو فراغ."""
    s = str(rec.sex or "").strip()
    if s.startswith("ذكر") or s.upper().startswith("M"):
        return "M"
    if s.startswith("أنثى") or s.startswith("انثى") or s.upper().startswith("F"):
        return "F"
    return ""


def nationality_code(rec: PassportData) -> str:
    """رمز الجنسية الثلاثي (SAU, ARE...) من الرمز أو من الاسم العربي."""
    code = str(rec.nationality or "").strip().upper()
    if len(code) == 3 and code.isalpha():
        return code
    return _AR_TO_CODE.get(str(rec.nationality_ar or "").strip(), "")


def travel_class_code(rec: PassportData) -> str:
    """J (أعمال/أولى) أو Y (سياحية) من درجة السفر، أو فراغ."""
    raw = str(rec.travel_class or "").strip()
    if not raw:
        return ""
    norm = re.sub(r"[ً-ْـ]", "", raw).replace("ة", "ه").lower()
    if any(w in norm for w in ("بزنس", "business", "اعمال", "رجال", "اولي",
                               "first", "درجه اولي")):
        return "J"
    if any(w in norm for w in ("سياحيه", "economy", "اقتصاديه", "eco", "y")):
        return "Y"
    upper = raw.upper()
    if upper in ("J", "C", "F"):
        return "J"
    if upper in ("Y", "M", "K", "B"):
        return "Y"
    return ""


def _to_ddmmmyy(iso: str) -> str:
    """يحوّل تاريخاً ISO إلى صيغة الطيران DDMMMYY (25FEB66)."""
    text = str(iso or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            d = datetime.strptime(text, fmt).date()
            return f"{d.day:02d}{_MONTHS[d.month - 1]}{d.year % 100:02d}"
        except ValueError:
            continue
    return text          # نُبقيه كما هو إن تعذّر الفهم


def _title(rec: PassportData) -> str:
    """لقب أماديوس: MR/MRS للبالغين، MSTR/MISS للأطفال دون 12."""
    g = gender_code(rec)
    child = False
    try:
        born = datetime.strptime(str(rec.birth_date).strip(), "%Y-%m-%d").date()
        today = date.today()
        age = today.year - born.year - ((today.month, today.day) < (born.month, born.day))
        child = age < 12
    except ValueError:
        child = False
    if g == "F":
        return "MISS" if child else "MRS"
    return "MSTR" if child else "MR"


def airline_row(rec: PassportData, serial: int) -> list:
    """صف كشف الطيران الواحد بالترتيب المطلوب."""
    last, first = split_name(rec)
    return [
        serial, last, first, str(rec.passport_number or "").strip().upper(),
        _to_ddmmmyy(rec.expiry_date), _to_ddmmmyy(rec.birth_date),
        gender_code(rec), nationality_code(rec), travel_class_code(rec),
        str(rec.family_number or "").strip(), str(rec.pnr or "").strip().upper(),
    ]


def airline_rows(records: list[PassportData]) -> list[list]:
    return [airline_row(rec, i) for i, rec in enumerate(records, start=1)]


def amadeus_name(rec: PassportData) -> str:
    """سطر اسم الراكب في أماديوس: NM1ALSHAMSI/ABDULLA MR (أو فراغ)."""
    last, first = split_name(rec)
    if not (last or first):
        return ""
    return f"NM1{last}/{first} {_title(rec)}".rstrip()


def amadeus_docs(rec: PassportData) -> str:
    """سطر وثيقة السفر DOCS في أماديوس (أو فراغ إن لا اسم)."""
    last, first = split_name(rec)
    if not (last or first):
        return ""
    nat = nationality_code(rec)
    return (
        f"SR DOCS HK1-P-{nat}-{str(rec.passport_number or '').strip().upper()}-"
        f"{nat}-{_to_ddmmmyy(rec.birth_date)}-{gender_code(rec)}-"
        f"{_to_ddmmmyy(rec.expiry_date)}-{last}-{first}"
    )


def amadeus_entry(rec: PassportData) -> str:
    """إدخال أماديوس كامل لراكب واحد (سطر الاسم + سطر DOCS)."""
    name = amadeus_name(rec)
    return f"{name}\n{amadeus_docs(rec)}" if name else ""


def amadeus_entries(records: list[PassportData]) -> str:
    """إدخالات أماديوس لكل الركّاب، مفصولة بسطر فارغ.

        NM1ALSHAMSI/ABDULLA MR
        SR DOCS HK1-P-ARE-AA0319030-ARE-27APR85-M-09JUL28-ALSHAMSI-ABDULLA
    """
    return "\n\n".join(e for e in (amadeus_entry(r) for r in records) if e)


# ------------------------------------------------------------------ التصدير
_HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# أعمدة الإكسل = كشف الطيران + عمودا إدخال أماديوس (اسم + DOCS) للنسخ منها
_EXCEL_COLUMNS = AIRLINE_COLUMNS + ("Amadeus Name", "Amadeus DOCS")
_EXCEL_WIDTHS = (5, 20, 22, 15, 11, 11, 8, 11, 12, 11, 12, 30, 62)


def export_airline_excel(records: list[PassportData], path: str | Path) -> Path:
    """يصدّر كشف الطيران إلى إكسل — إنجليزي LTR، مع عمودَي إدخال أماديوس."""
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Flight Manifest"
    ws.sheet_view.rightToLeft = False          # LTR

    ws.append(list(_EXCEL_COLUMNS))
    for col, width in enumerate(_EXCEL_WIDTHS, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[_col_letter(col)].width = width
    ws.row_dimensions[1].height = 28

    for rec, row in zip(records, airline_rows(records)):
        ws.append(row + [amadeus_name(rec), amadeus_docs(rec)])
        r = ws.max_row
        for col in range(1, len(_EXCEL_COLUMNS) + 1):
            cell = ws.cell(row=r, column=col)
            # أعمدة أماديوس محاذاة يسار (نص أوامر)، والبقية وسط
            align = "left" if col > len(AIRLINE_COLUMNS) else "center"
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = _BORDER

    ws.freeze_panes = "A2"
    if records:
        ws.auto_filter.ref = f"A1:{_col_letter(len(_EXCEL_COLUMNS))}{ws.max_row}"
    wb.save(path)
    return path


def _col_letter(index: int) -> str:
    """رقم عمود -> حرفه (يدعم ما بعد Z)."""
    letters = ""
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters

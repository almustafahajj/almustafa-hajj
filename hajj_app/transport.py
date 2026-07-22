"""كشف المواصلات: توزيع الحجّاج حسب وسيلة النقل (الباص) وتصديره إكسل."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .mrz import PassportData

# أعمدة كشف المواصلات (من اليمين لليسار في العرض العربي)
TRANSPORT_COLUMNS: tuple[str, ...] = (
    "م", "اسم الحاج", "الهاتف", "الفندق", "خدمة التنفيذي", "كرسي متحرك",
)
_WIDTHS = (5, 26, 14, 18, 14, 10)


def _wheelchair(rec: PassportData) -> str:
    return str(rec.wheelchair or "").strip()


def executive_display(rec: PassportData) -> str:
    """خدمة التنفيذي كما في الكشف العام — تُعرض القيمة كما هي (أو فارغة)."""
    return str(rec.executive_service or "").strip()


def distinct_transports(records: list[PassportData]) -> list[str]:
    """وسائل النقل الفريدة غير الفارغة، مرتّبة."""
    return sorted({str(r.transport or "").strip()
                   for r in records if str(r.transport or "").strip()})


def group_by_transport(
    records: list[PassportData],
) -> tuple[list[tuple[str, list[PassportData]]], list[PassportData]]:
    """يجمع الحجّاج حسب وسيلة النقل. يعيد (المجموعات المرتّبة، بلا مواصلات)."""
    groups: dict[str, list[PassportData]] = {}
    unassigned: list[PassportData] = []
    for rec in records:
        value = str(rec.transport or "").strip()
        if value:
            groups.setdefault(value, []).append(rec)
        else:
            unassigned.append(rec)
    ordered = [(value, groups[value]) for value in sorted(groups)]
    return ordered, unassigned


_HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_GROUP_FILL = PatternFill("solid", fgColor="3A342B")
_GROUP_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def export_transport_excel(records: list[PassportData], path: str | Path,
                           *, title: str = "كشف المواصلات") -> Path:
    """يصدّر كشف المواصلات إلى إكسل، مجموعاً بوسيلة النقل."""
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "المواصلات"
    ws.sheet_view.rightToLeft = True
    ncols = len(TRANSPORT_COLUMNS)

    ws.append([title] + [""] * (ncols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1)
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = Alignment(horizontal="center", vertical="center")

    ws.append(list(TRANSPORT_COLUMNS))
    for col, width in enumerate(_WIDTHS, start=1):
        cell = ws.cell(row=2, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 26

    def group_header(text: str) -> None:
        ws.append([text] + [""] * (ncols - 1))
        row = ws.max_row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1)
        cell.fill = _GROUP_FILL
        cell.font = _GROUP_FONT
        cell.alignment = Alignment(horizontal="right", vertical="center")

    def occupant(rec: PassportData, serial: int) -> None:
        ws.append([serial, rec.full_name_ar or rec.full_name_en or "—",
                   str(rec.phone or "").strip(), str(rec.hotel or "").strip(),
                   executive_display(rec), _wheelchair(rec)])
        row = ws.max_row
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

    groups, unassigned = group_by_transport(records)
    serial = 0
    for name, occ in groups:
        group_header(f"{name}  ({len(occ)})")
        for rec in occ:
            serial += 1
            occupant(rec, serial)
    if unassigned:
        group_header(f"بلا مواصلات  ({len(unassigned)})")
        for rec in unassigned:
            serial += 1
            occupant(rec, serial)

    ws.freeze_panes = "A3"
    wb.save(path)
    return path
